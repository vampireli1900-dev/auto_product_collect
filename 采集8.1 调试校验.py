import uiautomator2 as u2
import cv2
from ultralytics import YOLO
import time
import re
import requests
import easyocr
import pandas as pd
import os
import logging
import subprocess
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

logging.disable(logging.WARNING)
YOLO().verbose = False

# ====================== 初始化 ======================
d = u2.connect()
model = YOLO("runs/detect/pdd_logo_train-2/weights/best.pt")
subsidy_model = YOLO("runs/detect/subsidy_train/weights/best.pt")
detail_model = YOLO("runs/detect/product_detail_train/weights/best.pt")
reader = easyocr.Reader(['ch_sim'], gpu=False)

# ====================== 【新增配置项】 ======================
PRODUCT_LIST_FILE = "搜索名单.xlsx"
SEARCH_INTERVAL_SECONDS = 40
PACKAGE_NAME = "com.xunmeng.pinduoduo"

# ====================== 全局存储所有商品记录 ======================
record_list = []

# ====================== 最新表头 ======================
EXCEL_HEADER = [
    "序号",
    "货品名称",
    "关键词",
    "原价",
    "现价",
    "是否百亿补贴产品",
    "生产日期"
]

# ====================== 【新增调试用】校验记录存储与表头 ======================
debug_record_list = []
DEBUG_EXCEL_HEADER = [
    "校验序号",
    "搜索关键词",
    "提取到的商品标题",
    "搜索词匹配品牌",
    "商品标题匹配品牌",
    "品牌是否一致",
    "规格是否匹配通过",
    "品名匹配率(%)",
    "最终校验是否通过",
    "校验时间",
    "备注/失败原因"
]

# ====================== 品牌字典（你最新完整版） ======================
brand_lib = {
    "赫莲娜": ["HR", "HelenaRubinstein", "Helena Rubinstein", "赫莲娜"],
    "海蓝之谜": ["LM", "LaMer", "La Mer", "腊梅", "海蓝之谜"],
    "莱珀妮": ["LP", "LaPrairie", "La Prairie", "莱珀妮"],
    "希思黎": ["Sisley", "希思黎"],
    "法尔曼": ["Valmont", "法尔曼"],
    "兰蔻": ["Lancome", "Lancôme", "兰蔻"],
    "娇兰": ["Guerlain", "娇兰"],
    "倩碧": ["Clinique", "倩碧"],
    "娇韵诗": ["Clarins", "娇韵诗"],
    "科颜氏": ["Kiehls", "Kiehl's", "科颜氏"],
    "碧欧泉": ["Biotherm", "碧欧泉"],
    "薇姿": ["Vichy", "薇姿"],
    "德美乐嘉": ["Dermalogica", "德美乐嘉"],
    "雅诗兰黛": ["EsteeLauder", "Estée Lauder", "ESTEE LAUDER", "雅诗兰黛", "红石榴"],
    "大宝": ["Embryolisse", "大宝"],
    "薇迪薇奇": ["VidiVici", "Vidi Vici", "薇迪薇奇"],
    "肌肤之钥": ["CPB", "CleDePeauBeaute", "Cle de Peau Beauté", "肌肤之钥", "cledepeau"],
    "SK-II": ["SK2", "SKII", "SK-II"],
    "资生堂": ["Shiseido", "资生堂"],
    "安耐晒": ["安耐晒"],
    "黛珂": ["Decorte", "Decorté", "黛珂"],
    "城野医生": ["DrCiLabo", "Dr.Ci:Labo", "城野医生"],
    "茵芙莎": ["IPSA", "茵芙莎", "茵芙纱"],
    "宝丽": ["POLA", "宝丽"],
    "兰芝": ["Laneige", "兰芝"],
    "植村秀": ["ShuUemura", "Shu Uemura", "植村秀"],
    "汤姆福特": ["TF", "TomFord", "Tom Ford", "汤姆福特"],
    "圣罗兰": ["YSL", "YvesSaintLaurent", "Yves Saint Laurent", "圣罗兰"],
    "魅可": ["MAC", "M.A.C", "魅可"],
    "纳斯": ["NARS", "纳斯"],
    "芭比布朗": ["BobbiBrown", "Bobbi Brown", "BB", "芭比布朗", "芭比波朗"],
    "纪梵希": ["Givenchy", "纪梵希"],
    "苏秘": ["Sum37", "Su:m37", "苏秘", "苏秘37"],
    "衰败城市": ["UrbanDecay", "Urban Decay", "衰败城市"],
    "罗拉": ["LauraMercier", "Laura Mercier", "罗拉"],
    "后": ["WHOO", "TheHistoryOfWhoo", "The History Of Whoo", "后", "拱辰享"],
    "雪花秀": ["Sulwhasoo", "雪花", "后雪", "雪花秀"],
    "祖玛珑": ["JM", "JoMalone", "Jo Malone", "祖马龙", "祖玛珑"],
    "芦丹氏": ["SL", "SergeLutens", "Serge Lutens", "芦丹氏"],
    "百瑞德": ["Byredo", "百瑞德"],
    "爱马仕": ["Hermes", "Hermès", "爱马仕"],
    "古驰": ["Gucci", "古驰"],
    "范思哲": ["Versace", "范思哲"],
    "阿玛尼": ["Armani", "阿玛尼"],
    "巴宝莉": ["Burberry", "巴宝莉", "博柏利"],
    "蔻依": ["Chloe", "Chloé", "克洛伊", "蔻依"],
    "莫杰": ["MarcJacobs", "Marc Jacobs", "MJ", "莫杰"],
    "纳西索": ["Narciso", "纳西索"],
    "帕尔玛之水": ["AcquaDiParma", "Acqua di Parma", "帕尔玛之水"],
    "梅森马吉拉": ["MaisonMargiela", "Maison Margiela", "MM", "梅森马吉拉", "马丁马吉拉"],
    "欧珑": ["AtelierCologne", "Atelier Cologne", "欧珑"],
    "宝格丽": ["Bvlgari", "宝格丽"],
    "杜鲁萨迪": ["Trussardi", "杜鲁萨迪"],
    "卡尔文克雷恩": ["CK", "CalvinKlein", "Calvin Klein", "卡尔文克雷恩"],
    "缪缪": ["MiuMiu", "miumiu", "Miu Miu", "缪缪"],
    "香奈儿": ["Chanel", "香奈儿"],
    "卡诗": ["Kerastase", "Kérastase", "卡诗"],
    "欧舒丹": ["Loccitane", "L'Occitane", "欧舒丹"],
    "欧莱雅": ["Loreal", "L'Oréal", "欧莱雅"],
    "潘婷": ["Pantene", "潘婷"],
    "大卫杜夫": ["Davidoff", "大卫杜夫"],
    "拉夫劳伦": ["RalphLauren", "Ralph Lauren", "拉夫劳伦"],
    "馥蕾诗": ["Fresh", "馥蕾诗"],
    "伟博": ["Webber", "伟博"],
    "慕拉得": ["Murad", "慕拉得"],
    "未来驱蚊": ["VAPE", "未來", "未来驅蚊"],
    "澳洲NatureBOBO": ["NatureBOBO", "Nature BOBO", "澳洲"],
    "旧街场": ["OldTown", "Old Town", "旧街场"],
    "费列罗": ["Ferrero", "费列", "费列罗"]
}


# ====================== 品牌匹配（大小写不敏感） ======================
def match_brand(text):
    text_lower = str(text).lower()
    matched_brand = "未匹配"
    for standard_name, alias_list in brand_lib.items():
        for alias in alias_list:
            alias_lower = alias.lower()
            if alias_lower in text_lower:
                return standard_name
    return matched_brand


# ====================== 规格+型号提取 ======================
def extract_specs(text):
    text = str(text).strip()
    temp = text
    res = []
    rules = [
        r'[A-Za-z0-9]+#', r'#[A-Za-z0-9]+',
        r'[A-Za-z]+[0-9]+[A-Za-z]*', r'(?<!\.)\d+[A-Za-z]+',
        r'\d+色', r'\d+号', r'\d+\.\d+\s*[mlgMLG]+', r'\d+\s*[mlgMLG]+',
        r'EDT|EDP|浓香|淡香',
        r'对装|两支装|三支装|两瓶装|双支装|\*2|x2|X2',
        r'新款|新版|旧版|经典款', r'\d+款|\d+年',
        r'(?<![A-Za-z])\d+\.?\d*(?![A-Za-z#])',
    ]
    for p in rules:
        ms = re.findall(p, temp, re.I)
        for v in ms:
            if v and v not in res:
                res.append(v)
                temp = temp.replace(v, " ")
    return res


# ====================== 品名清洗（删除品牌+规格） ======================
def clean_name(text, brand, specs):
    s = str(text).strip()
    if brand != "未匹配":
        for alias in brand_lib[brand]:
            s = re.compile(re.escape(alias), re.I).sub("", s)
    for sp in specs:
        s = re.compile(re.escape(sp), re.I).sub("", s)
    s = re.sub(r'[^\w]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


# ====================== 字符匹配率 >=60% ======================
def char_match_rate(a, b):
    a = re.sub(r'[^a-z0-9\u4e00-\u9fff]', '', a.lower())
    b = re.sub(r'[^a-z0-9\u4e00-\u9fff]', '', b.lower())
    if not a or not b:
        return 0.0
    set_a, set_b = set(a), set(b)
    common = len(set_a & set_b)
    return common / len(set_a)


# ====================== 🔥 最终四重校验函数（接入你业务+新增调试记录） ======================
def validate_product(search_word, product_title):
    # 1. 提取搜索词信息
    s_brand = match_brand(search_word)
    s_spec = extract_specs(search_word)
    s_name = clean_name(search_word, s_brand, s_spec)

    # 2. 提取商品信息
    p_brand = match_brand(product_title)
    p_spec = extract_specs(product_title)
    p_name = clean_name(product_title, p_brand, p_spec)

    # 3. 品牌必须一致
    brand_ok = (s_brand == p_brand) if s_brand != "未匹配" else True

    # 4. 规格必须包含（套装/版本不敏感）
    spec_ok = True
    ignore = {"对装", "两支装", "三支装", "新版", "旧版", "经典款", "新款"}
    for sp in s_spec:
        if any(i in sp for i in ignore):
            continue
        if not any(sp.lower() in ps.lower() for ps in p_spec):
            spec_ok = False
            break

    # 5. 品名匹配率 ≥60%
    rate = char_match_rate(s_name, p_name)
    name_ok = rate >= 0.6

    final = brand_ok and spec_ok and name_ok

    # 【新增调试记录】生成失败原因备注
    fail_reason = []
    if not brand_ok:
        fail_reason.append(f"品牌不一致：搜索词匹配{s_brand}，商品匹配{p_brand}")
    if not spec_ok:
        fail_reason.append(f"规格不匹配：搜索词规格{s_spec}，商品规格{p_spec}")
    if not name_ok:
        fail_reason.append(f"品名匹配率不足：{rate:.1%} < 60%")
    remark = "；".join(fail_reason) if fail_reason else "校验通过"

    # 【新增调试记录】追加到全局列表
    debug_record_list.append([
        len(debug_record_list) + 1,  # 校验序号
        search_word,
        product_title,
        s_brand,
        p_brand,
        "是" if brand_ok else "否",
        "是" if spec_ok else "否",
        round(rate * 100, 2),
        "是" if final else "否",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        remark
    ])

    # 打印调试（你原来的风格）
    print(f"[校验] 品牌={s_brand}/{p_brand} → {brand_ok}")
    print(f"[校验] 规格匹配 → {spec_ok}")
    print(f"[校验] 品名匹配率={rate:.1%} → {name_ok}")
    print(f"[校验] 最终通过={final}")
    return final


# ====================== 【修复版】保存校验记录Excel 支持追加不覆盖 ======================
def save_debug_excel():
    if not debug_record_list:
        print("⚠️ 暂无校验调试数据，跳过保存")
        return

    file_path = "商品校验调试记录.xlsx"
    new_df = pd.DataFrame(debug_record_list, columns=DEBUG_EXCEL_HEADER)

    # 如果文件已存在，读取旧数据 拼接追加
    if os.path.exists(file_path):
        try:
            old_df = pd.read_excel(file_path)
            new_df = pd.concat([old_df, new_df], ignore_index=True)
        except:
            # 读失败就覆盖新建，防止卡死
            pass

    # 去重：按 搜索关键词+商品标题+校验时间 防重复
    new_df.drop_duplicates(
        subset=["搜索关键词", "提取到的商品标题", "校验时间"],
        keep="last",
        inplace=True
    )

    # 重新编排序号
    new_df["校验序号"] = list(range(1, len(new_df) + 1))

    # 保存并美化格式
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        new_df.to_excel(writer, index=False, sheet_name='校验记录')
        workbook = writer.book
        worksheet = writer.sheets['校验记录']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        data_align = Alignment(vertical="center", wrap_text=True)
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = data_align
                if col == 8:
                    cell.number_format = '0.00"%"'

        for col in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    current_length = len(str(cell.value))
                    if current_length > max_length:
                        max_length = current_length
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width

    print(f"\n📋 已追加保存校验调试记录至：{file_path}，当前总校验记录数：{len(new_df)}")


# ====================== 保存Excel（主表+新增调试表同步保存） ======================
def save_all_to_excel():
    if not record_list:
        print("⚠️ 暂无采集数据，跳过保存")
        return
    # 保存主采集表
    df = pd.DataFrame(record_list, columns=EXCEL_HEADER)
    file = "商品采集汇总.xlsx"
    if os.path.exists(file):
        old_df = pd.read_excel(file)
        df = pd.concat([old_df, df], ignore_index=True)
    df = df.drop_duplicates(subset=["序号", "货品名称", "现价"], keep="last")
    df.to_excel(file, index=False)
    print(f"\n📁 已批量保存全部采集记录至：{file}，当前总行数：{len(df)}")

    # 【新增】同步保存调试校验表
    save_debug_excel()


# ====================== 回到首页 ======================
def go_to_pinduoduo_home():
    print("\n🔴 检测到异常，正在强制跳回拼多多首页...")
    try:
        subprocess.run([
            "adb", "shell", "am", "start", "-S", "-n",
            "com.xunmeng.pinduoduo/com.xunmeng.pinduoduo.ui.activity.MainFrameActivity"
        ], check=True, timeout=15)
        time.sleep(6)
        print("🟢 已回到拼多多首页")
    except:
        subprocess.run(["adb", "shell", "am", "force-stop", PACKAGE_NAME])
        time.sleep(2)
        subprocess.run(["adb", "shell", "am", "start", "-n",
                        "com.xunmeng.pinduoduo/com.xunmeng.pinduoduo.ui.activity.MainFrameActivity"])
        time.sleep(8)


# ====================== 读取名单 ======================
def load_product_list_with_status():
    if not os.path.exists(PRODUCT_LIST_FILE):
        print(f"❌ 未找到名单文件")
        return []
    df = pd.read_excel(PRODUCT_LIST_FILE)
    if "状态" not in df.columns:
        df["状态"] = "未采集"
    if "序号" not in df.columns:
        df["序号"] = range(1, len(df)+1)
        df.to_excel(PRODUCT_LIST_FILE, index=False)
    todo = df[df["状态"] == "未采集"].copy()
    print(f"✅ 总商品数：{len(df)} | 待采集：{len(todo)}")
    return todo

def mark_product_as_done_by_index(index_num):
    """用序号标记，100% 不会因为商品名不匹配失败"""
    try:
        df = pd.read_excel(PRODUCT_LIST_FILE)
        # 用序号定位，永远不会错
        df.loc[df["序号"] == index_num, "状态"] = "已采集"
        df.to_excel(PRODUCT_LIST_FILE, index=False)
        print(f"✅ 已标记序号 {index_num} 为已采集")
    except Exception as e:
        print(f"❌ 标记状态失败：{str(e)}")
        # 备份损坏的文件
        backup_name = f"{PRODUCT_LIST_FILE}.bak"
        os.rename(PRODUCT_LIST_FILE, backup_name)
        print(f"⚠️ 已将原文件备份为 {backup_name}")


# ====================== 你原来的AI函数（保留但不做主校验） ======================
def is_same_product_by_llm(search_word, product_title):
    return validate_product(search_word, product_title)


# ====================== 以下所有代码完全保留你原版 ======================
def search_product(keyword):
    print(f"\n🔍 开始搜索商品：{keyword}")
    width, height = d.window_size()
    search_y = int(height * 200 / 2400)
    d.swipe(width // 2, height // 2, width // 2, height // 2 + 400)
    time.sleep(1)
    d.click(width // 2, search_y)
    time.sleep(0.8)
    d(className="android.widget.EditText").clear_text()
    d(className="android.widget.EditText").set_text(keyword)
    time.sleep(0.8)
    d(text="搜索", className="android.widget.TextView").click()
    print("✅ 搜索完成，等待列表加载")
    time.sleep(3)


def scan_list_products():
    d.screenshot("list_screen.jpg")
    img = cv2.imread("list_screen.jpg")
    results = model(img, conf=0.2)
    products = []
    for r in results:
        boxes = r.boxes
        for box in boxes:
            cls = int(box.cls)
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
            if cls == 0:
                products.append({"type": "baiyi", "cx": cx, "cy": cy})
            elif cls == 1:
                products.append({"type": "brand", "cx": cx, "cy": cy})
            elif cls == 2:
                products.append({"type": "global", "cx": cx, "cy": cy})
    return products


def get_products_with_tags():
    products = scan_list_products()
    grouped = {}
    for p in products:
        cx, cy = p["cx"], p["cy"]
        matched = False
        for key in list(grouped.keys()):
            ecx, ecy = key
            if abs(cx - ecx) < 200 and abs(cy - ecy) < 180:
                grouped[key]["tags"].add(p["type"])
                matched = True
                break
        if not matched:
            grouped[(cx, cy)] = {"tags": {p["type"]}, "cx": cx, "cy": cy}
    return list(grouped.values())


def get_priority(tags):
    if "baiyi" in tags and "brand" in tags:
        return 4
    elif "baiyi" in tags:
        return 3
    elif "brand" in tags:
        return 2
    elif "global" in tags:
        return 1
    else:
        return 0


def is_all_global(item_list):
    for i in item_list:
        if "baiyi" in i["tags"] or "brand" in i["tags"]:
            return False
    return True


def scroll_down_once():
    d.swipe(500, 1800, 500, 600, 0.3)
    time.sleep(2.5)


def scroll_to_top():
    for _ in range(2):
        d.swipe(500, 500, 500, 1900, 0.3)
        time.sleep(0.8)


def sort_products_by_priority():
    raw = get_products_with_tags()
    if is_all_global(raw):
        scroll_down_once()
        raw = get_products_with_tags()
        if is_all_global(raw):
            scroll_down_once()
            raw = get_products_with_tags()
            if is_all_global(raw):
                scroll_to_top()
                raw = get_products_with_tags()
    s = sorted(raw, key=lambda x: get_priority(x["tags"]), reverse=True)
    if is_all_global(s):
        s = s[:2]
    return s


def is_subsidy_product():
    return "百亿补贴" in d.dump_hierarchy() or "官方补贴" in d.dump_hierarchy()


def extract_product_info(xml_content, search_word):
    import re
    def get_ngram_pairs(text, n=2):
        text = re.sub(r'[^a-z0-9\u4e00-\u9fff]', '', text.lower())
        return [text[i:i + n] for i in range(len(text) - n + 1)] if len(text) >= n else [text]

    def get_single_chars(text):
        text = re.sub(r'[^a-z0-9\u4e00-\u9fff]', '', text.lower())
        return [c for c in text]

    def count_chinese(text):
        return len(re.findall(r'[\u4e00-\u9fff]', text))

    search_cn_count = count_chinese(search_word)
    desc_list = re.findall(r'content-desc="([^"]+)"', xml_content)

    best_title = ""
    best_count = 0

    blacklist = [
        "电池", "状态栏", "电量", "百分之", "WLAN", "手机信号", "5G", "4G",
        "通知", "高德", "淘宝", "浏览器", "手机管家", "振铃器", "静音",
        "返回", "分享", "店铺", "收藏", "客服", "工具栏", "顶部", "拼小圈",
        "¥", "￥", "大促价", "已抢", "假一赔十", "100%正品", "拼单价",
        "狂降", "直接成团", "买过", "次", "图片", "该店", "tronplayer_view", "查看全部"
    ]

    search_pairs = get_ngram_pairs(search_word)
    for desc in desc_list:
        desc = desc.strip()
        if any(kw in desc for kw in blacklist):
            continue
        desc_clean = re.sub(r'[^a-z0-9\u4e00-\u9fff]', '', desc.lower())
        match_count = sum(1 for p in search_pairs if p in desc_clean)
        if match_count > best_count and match_count > 0:
            best_count = match_count
            best_title = desc
        elif match_count == best_count and match_count > 0:
            if len(desc) > len(best_title):
                best_title = desc

    if not best_title:
        search_chars = get_single_chars(search_word)
        best_count = 0
        for desc in desc_list:
            desc = desc.strip()
            if any(kw in desc for kw in blacklist):
                continue
            if count_chinese(desc) < search_cn_count:
                continue
            desc_clean = re.sub(r'[^a-z0-9\u4e00-\u9fff]', '', desc.lower())
            match_count = sum(1 for c in search_chars if c in desc_clean)
            if match_count > best_count and match_count > 0:
                best_count = match_count
                best_title = desc
            elif match_count == best_count and match_count > 0:
                if len(desc) > len(best_title):
                    best_title = desc

    price_pattern = r'[¥￥]\s*(\d+(?:\.\d+)?)'
    all_prices = re.findall(price_pattern, xml_content)
    price_nums = [float(p) for p in all_prices]

    original_price = None
    current_price = None

    if len(price_nums) > 0:
        prices = sorted(list(set(price_nums)))
        valid = []
        for i in prices:
            keep = True
            for j in prices:
                if i == j:
                    continue
                if max(i, j) >= min(i, j) * 10:
                    s_i = str(int(round(i)))
                    s_j = str(int(round(j)))
                    if len(s_i) >= 3 and len(s_j) >= 3 and s_i[:3] == s_j[:3]:
                        if i > j:
                            keep = False
                        break
            if keep:
                valid.append(i)
        if valid:
            current_price = str(min(valid))
            original_price = str(max(valid))

    return {
        "title": best_title.strip() if best_title else "",
        "original_price": original_price,
        "current_price": current_price
    }


def find_and_click_detail(max_scroll=7):
    for _ in range(max_scroll):
        img = d.screenshot(format="opencv")
        res = detail_model(img, conf=0.75)
        box = None

        for r in res:
            for b in r.boxes:
                if int(b.cls[0]) == 0:
                    box = tuple(map(int, b.xyxy[0]))
                    break
            if box:
                break

        if box:
            x1, y1, x2, y2 = box
            crop_img = img[y1:y2, x1:x2]

            ocr_result = reader.readtext(crop_img)
            full_text = ""
            for item in ocr_result:
                full_text += item[1]

            if "商品详情" in full_text:
                cx = (x1 + x2) // 2
                cy = (y1 + y2) // 2
                d.click(cx, cy)
                time.sleep(1.5)
                page_xml = d.dump_hierarchy()
                if "生产日期" in page_xml:
                    return True
                else:
                    d.press("back")
                    return False

        d.swipe(500, 1800, 500, 600, 0.25)
        time.sleep(0.8)

    return False


def get_production_date_from_xml(xml):
    m = re.search(r'text="(\d{4}-\d{1,2}-\d{1,2})"', xml)
    return m.group(1) if m else None


def get_date_with_retry():
    return get_production_date_from_xml(d.dump_hierarchy())


def collect_single_product(search_word, serial_num):
    xml = d.dump_hierarchy()
    info = extract_product_info(xml, search_word)
    title = info["title"]
    ori = info["original_price"]
    cur = info["current_price"]
    subsidy = "是" if is_subsidy_product() else "否"
    detail = find_and_click_detail()
    date = get_date_with_retry() if detail else ""

    # ✅ 调用新校验
    match_pass = validate_product(search_word, title)

    record_list.append([serial_num, title, search_word, ori, cur, subsidy, date])
    print("\n" + "=" * 80)
    print(f"货品名称：{title}")
    print(f"关键词：{search_word}")
    print(f"原价：{ori} | 现价：{cur}")
    print(f"百亿补贴：{subsidy} | 日期：{date}")
    print(f"校验通过：{match_pass}")
    print("=" * 80)
    return {"title": title, "subsidy": subsidy, "date": date, "found_detail": detail}


def select_and_collect_best_product(search_word, serial_num):
    sorted_prods = sort_products_by_priority()
    if not sorted_prods:
        print("❌ 未识别商品")
        return None
    hp = get_priority(sorted_prods[0]["tags"])
    candidates = [p for p in sorted_prods if get_priority(p["tags"]) == hp]
    time.sleep(0.5)
    for i, p in enumerate(candidates):
        print(f"\n--- 进入商品 {i + 1}/{len(candidates)} ---")
        d.click(p["cx"], p["cy"])
        time.sleep(1)
        res = collect_single_product(search_word, serial_num)
        if res["found_detail"]:
            d.press("back")
            time.sleep(1.5)
            d.press("back")
        else:
            d.press("back")
        time.sleep(1)
    return True


# ====================== 主循环关键修改 ======================
def main():
    print("🚀 启动采集 + 品牌规格品名四重校验 + 60%匹配率 + 调试校验记录")
    todo = load_product_list_with_status()
    if todo.empty:
        print("🎉 全部完成")
        return
    for _, row in todo.iterrows():
        kw = str(row["货品名称"]).strip()
        idx = int(row["序号"])  # 拿序号
        ok = False
        while not ok:
            try:
                search_product(kw)
                select_and_collect_best_product(kw, idx)
                save_all_to_excel()          # 先保存，打印日志
                mark_product_as_done_by_index(idx)  # 用序号标记，永不失效
                time.sleep(SEARCH_INTERVAL_SECONDS)
                ok = True
            except Exception as e:
                print(f"❌ 异常：{e}")
                go_to_pinduoduo_home()
    print("🎉 全部采集完成！")


if __name__ == "__main__":
    main()