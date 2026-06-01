import pandas as pd
import numpy as np
import re
import os
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ================== 配置 ==================
INPUT_FILE = "录入表.xlsx"
COLLECT_FILE = "商品采集汇总.xlsx"
OUTPUT_FILE = "录入表_更新.xlsx"
# 基准日期：1900-01-01
BASE_DATE = pd.Timestamp("1900-01-01")

QUANTITY_PATTERN = re.compile(r'(双支|两支|2支|\*2|对装|双只|两瓶|两支装|双支装)', re.IGNORECASE)

# ================== 辅助函数 ==================
def normalize_price(value):
    if pd.isna(value):
        return np.nan
    if isinstance(value, str):
        value = value.strip().replace('¥', '').replace('元', '')
        if value == '' or value == '-':
            return np.nan
    try:
        return float(value)
    except:
        return np.nan

def parse_date(date_val):
    """解析日期，无效或早于1900年返回NaT"""
    if pd.isna(date_val):
        return pd.NaT
    try:
        dt = pd.to_datetime(date_val, errors='coerce')
        if pd.notna(dt) and dt.year < 1900:
            return pd.NaT
        return dt
    except:
        return pd.NaT

def get_shelf_days_from_expiry(expiry_date):
    """
    如果失效日期是早期（年份<=1905），则视为保质期长度 = expiry_date - BASE_DATE。
    否则返回 None（表示需要从生产日期到失效日期计算）。
    """
    exp = parse_date(expiry_date)
    if pd.isna(exp):
        return None
    if exp.year <= 1905:
        days = (exp - BASE_DATE).days
        # 防止负数或异常值（如差0天）
        if days > 0 and days < 3650:  # 最多10年
            return days
    return None

def get_shelf_life_days(prod_date, exp_date):
    """
    根据生产日期和失效日期计算保质期天数。
    优先使用早期失效日期隐含的保质期，否则计算实际差值。
    返回天数或 None。
    """
    # 先尝试从失效日期直接获取保质期（早期日期）
    shelf_days = get_shelf_days_from_expiry(exp_date)
    if shelf_days is not None:
        return shelf_days

    # 正常情况：计算生产到失效的天数
    prod = parse_date(prod_date)
    exp = parse_date(exp_date)
    if pd.isna(prod) or pd.isna(exp):
        return None
    delta = (exp - prod).days
    if 0 < delta <= 3650:
        return delta
    return None

def adjust_price_by_quantity(price, keyword, product_name):
    if pd.isna(price):
        return price
    kw_has = bool(QUANTITY_PATTERN.search(str(keyword))) if not pd.isna(keyword) else False
    name_has = bool(QUANTITY_PATTERN.search(str(product_name))) if not pd.isna(product_name) else False
    if not kw_has and name_has:
        return price / 2.0
    elif kw_has and not name_has:
        return price * 2.0
    else:
        return price

def compute_candidate_price(row):
    spec_price = normalize_price(row.get('规格价格', np.nan))
    if not pd.isna(spec_price):
        return spec_price
    is_bai = str(row.get('是否百亿补贴产品', '')).strip() == '是'
    orig_price = normalize_price(row.get('原价', np.nan))
    curr_price = normalize_price(row.get('现价', np.nan))
    if is_bai:
        return orig_price if not pd.isna(orig_price) else curr_price
    else:
        return curr_price if not pd.isna(curr_price) else orig_price

def select_best_record(records):
    if records.empty:
        return None
    records = records.copy()
    records['_base_price'] = records.apply(compute_candidate_price, axis=1)
    records['_final_price'] = records.apply(
        lambda row: adjust_price_by_quantity(row['_base_price'], row.get('关键词'), row.get('货品名称')),
        axis=1
    )
    records['_parsed_date'] = records['生产日期'].apply(parse_date)
    records['_has_date'] = records['_parsed_date'].notna()
    records['_bai_score'] = (records['是否百亿补贴产品'].astype(str).str.strip() == '是').astype(int)
    records_sorted = records.sort_values(
        by=['_bai_score', '_has_date', '_parsed_date', '_final_price'],
        ascending=[False, False, False, True],
        na_position='last'
    )
    return records_sorted.iloc[0]

def apply_excel_style(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        data_align = Alignment(vertical="center", wrap_text=True)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).alignment = data_align
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_len = max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)), default=10)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        wb.save(file_path)
    except Exception as e:
        print(f"应用样式时出错: {e}")

# ================== 主流程 ==================
def main():
    # 读取文件
    df_input = pd.read_excel(INPUT_FILE, dtype=str)
    df_collect = pd.read_excel(COLLECT_FILE, dtype=str)

    df_input.columns = df_input.columns.str.strip()
    df_collect.columns = df_collect.columns.str.strip()

    if '序号' not in df_input.columns:
        print("录入表中缺少'序号'列")
        return
    df_input['序号'] = df_input['序号'].astype(str)

    required_cols = ['序号', '校验通过', '是否百亿补贴产品', '生产日期', '原价', '现价', '规格价格', '货品名称', '关键词']
    for col in required_cols:
        if col not in df_collect.columns:
            print(f"采集汇总表缺少必要列: {col}")
            return

    # 筛选校验通过的行
    df_collect_valid = df_collect[df_collect['校验通过'] == '✅'].copy()
    if df_collect_valid.empty:
        print("没有校验通过的商品，录入表不会更新")
        df_input.to_excel(OUTPUT_FILE, index=False)
        apply_excel_style(OUTPUT_FILE)
        return

    df_collect_valid['序号'] = df_collect_valid['序号'].astype(str)

    # ---------- 构建保质期映射（从录入表原始数据） ----------
    shelf_life_map = {}  # 序号 -> 保质期天数
    for idx, row in df_input.iterrows():
        seq = str(row['序号'])
        prod_orig = row.get('生产日期', '')
        exp_orig = row.get('失效日期', '')
        days = get_shelf_life_days(prod_orig, exp_orig)
        if days is not None:
            shelf_life_map[seq] = days
        # 如果当前行没有有效保质期，暂时不设置（后面可能用默认值）

    # ---------- 分组选择最佳采集记录 ----------
    grouped = df_collect_valid.groupby('序号')
    update_map = {}  # 序号 -> {价格, 生产日期, 是否百亿补贴产品, 失效日期}

    for seq, group in grouped:
        best = select_best_record(group)
        if best is None:
            continue

        # 计算最终价格
        raw_price = compute_candidate_price(best)
        final_price = adjust_price_by_quantity(raw_price, best.get('关键词'), best.get('货品名称'))

        # 新生产日期
        new_prod_str = best.get('生产日期')
        new_prod_dt = parse_date(new_prod_str)

        # 新失效日期
        new_exp_date = ''
        if pd.notna(new_prod_dt):
            # 获取保质期天数：优先用原表的映射，否则使用默认（例如1095天=3年）
            shelf_days = shelf_life_map.get(seq)
            if shelf_days is None:
                # 默认3年
                shelf_days = 1095
            new_exp_dt = new_prod_dt + timedelta(days=shelf_days)
            new_exp_date = new_exp_dt.strftime('%Y/%m/%d')
        else:
            new_exp_date = ''  # 生产日期无效则失效日期留空

        # 格式化生产日期为 YYYY/MM/DD
        formatted_prod = ''
        if pd.notna(new_prod_dt):
            formatted_prod = new_prod_dt.strftime('%Y/%m/%d')

        is_bai = best.get('是否百亿补贴产品', '')

        update_map[seq] = {
            '价格': final_price,
            '生产日期': formatted_prod,
            '是否百亿补贴产品': is_bai,
            '失效日期': new_exp_date
        }

    # 更新录入表
    if '价格' in df_input.columns:
        df_input['价格'] = df_input['价格'].apply(normalize_price)
    else:
        df_input['价格'] = np.nan

    for idx, row in df_input.iterrows():
        seq = str(row['序号'])
        if seq in update_map:
            updates = update_map[seq]
            df_input.at[idx, '价格'] = updates['价格']
            df_input.at[idx, '生产日期'] = updates['生产日期']
            df_input.at[idx, '是否百亿补贴产品'] = updates['是否百亿补贴产品']
            df_input.at[idx, '失效日期'] = updates['失效日期']

    # 价格保留两位小数
    if '价格' in df_input.columns:
        df_input['价格'] = df_input['价格'].round(2)

    df_input.to_excel(OUTPUT_FILE, index=False)
    apply_excel_style(OUTPUT_FILE)
    print(f"处理完成，结果已保存至: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()