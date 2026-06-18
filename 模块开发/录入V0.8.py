import pandas as pd
import numpy as np
import re
import os
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ================== 配置 ==================
INPUT_FILE = "录入表.xlsx"
COLLECT_FILE = "商品采集汇总.xlsx"
OUTPUT_FILE = "录入表_更新.xlsx"
BASE_DATE = pd.Timestamp("1900-01-01")

# 数量词正则（包含双瓶装等）
QUANTITY_PATTERN = re.compile(
    r'(双支|两支|2支|\*2|对装|双只|两瓶|2瓶|两支装|双支装|双瓶装|两瓶装|双包装|两份|双份|两只装|2只装|两只|2只)',
    re.IGNORECASE
)
"""
从同一序号的多条采集记录中选出最佳一条。

优先级规则（从高到低）：
1. 校验通过 —— 已在调用前过滤，所有记录均满足“校验通过 == ✅”，因此不在此处重复排序。
2. 规格匹配优先级 —— 按“匹配规格”列的值：
    - 匹配成功（值非空且不包含“匹配失败”，如“已选: xxx”） 优先级最高（2）
    - 空值或缺失（无规格信息）                            优先级中等（1）
    - 明确为“匹配失败”                                  优先级最低（0）
3. 是否百亿补贴 —— “是” > “否”
4. 是否有生产日期 —— 有生产日期 > 无生产日期
5. 生产日期早晚 —— 越晚（日期越大）越好（降序）
6. 最终价格 —— 经数量调整（双份乘/除）后，价格越低越好（升序）

排序时依次按上述维度降序/升序，前者维度相同时才比较后者。
"""
# ================== 辅助函数 ==================
def normalize_price(value):
    if pd.isna(value):
        return np.nan
    if isinstance(value, str):
        value = value.strip().replace('¥', '').replace('元', '')
        if value == '' or value == '-':
            return np.nan
    try:
        return float(value)
    except:
        return np.nan


def parse_date(date_val):
    if pd.isna(date_val):
        return pd.NaT
    try:
        dt = pd.to_datetime(date_val, errors='coerce')
        if pd.notna(dt) and dt.year < 1900:
            return pd.NaT
        return dt
    except:
        return pd.NaT


def get_shelf_days_from_expiry(expiry_date):
    exp = parse_date(expiry_date)
    if pd.isna(exp):
        return None
    if exp.year <= 1905:
        days = (exp - BASE_DATE).days
        if days > 0 and days < 3650:
            return days
    return None


def get_shelf_life_days(prod_date, exp_date):
    shelf_days = get_shelf_days_from_expiry(exp_date)
    if shelf_days is not None:
        return shelf_days
    prod = parse_date(prod_date)
    exp = parse_date(exp_date)
    if pd.isna(prod) or pd.isna(exp):
        return None
    delta = (exp - prod).days
    if 0 < delta <= 3650:
        return delta
    return None


def adjust_price_by_quantity(price, keyword, product_name):
    if pd.isna(price):
        return price, ''
    kw_has = bool(QUANTITY_PATTERN.search(str(keyword))) if not pd.isna(keyword) else False
    name_has = bool(QUANTITY_PATTERN.search(str(product_name))) if not pd.isna(product_name) else False
    if not kw_has and name_has:
        return price / 2.0, '价格减半（双份）'
    elif kw_has and not name_has:
        return price * 2.0, '价格翻倍（双份）'
    else:
        return price, ''


def compute_candidate_price(row):
    spec_price = normalize_price(row.get('规格价格', np.nan))
    is_bai = str(row.get('是否百亿补贴产品', '')).strip() == '是'
    orig_price = normalize_price(row.get('原价', np.nan))
    curr_price = normalize_price(row.get('现价', np.nan))

    # 特殊规则：百亿补贴 + 规格价存在 + 与现价相差 < 50 → 优先用原价
    if is_bai and not pd.isna(spec_price) and not pd.isna(curr_price):
        if abs(spec_price - curr_price) < 50:
            if not pd.isna(orig_price):
                return orig_price
            # 原价无效时，不返回规格价？按需求描述仅当相近时选原价，若无原价则可能仍需规格价，但为避免错乱，这里继续执行后续逻辑
            # 即如果原价无效，就跳过后面的规格价优先，但原规则仍可能用规格价，我们保留后续判断

    # 原有优先级：规格价 > 百亿补贴?原价:现价
    if not pd.isna(spec_price):
        return spec_price
    if is_bai:
        return orig_price if not pd.isna(orig_price) else curr_price
    else:
        return curr_price if not pd.isna(curr_price) else orig_price


def select_best_record(records, last_price_map=None):
    """
    从同一序号的多条采集记录中选出最佳一条。
    增加异常价格修正：当原价和现价相差>8倍时，根据上次价格修正。
    """
    if records.empty:
        return None
    records = records.copy()

    # ----- 价格修正函数（内部） -----
    def fix_price_anomaly(row, last_price):
        orig = normalize_price(row.get('原价'))
        curr = normalize_price(row.get('现价'))
        if pd.isna(orig) or pd.isna(curr):
            return row
        # 检查倍数是否大于8
        max_val = max(orig, curr)
        min_val = min(orig, curr)
        if min_val == 0:
            return row
        if max_val / min_val <= 8:
            return row
        # 如果上次价格无效，无法修正
        if pd.isna(last_price) or last_price == 0:
            return row
        # 选择与上次价格更接近的作为真实价格
        dist_orig = abs(orig - last_price)
        dist_curr = abs(curr - last_price)
        if dist_orig < dist_curr:
            real_price = orig
            fake_price = curr
        else:
            real_price = curr
            fake_price = orig
        # 检查前两位数字是否相同（去除小数点）
        str_real = str(real_price).replace('.', '')
        str_fake = str(fake_price).replace('.', '')
        if len(str_real) >= 2 and len(str_fake) >= 2 and str_real[:2] == str_fake[:2]:
            # 修正错误的价格为真实价格
            if real_price == orig:
                # 原价正确，修正现价
                row['现价'] = real_price
            else:
                # 现价正确，修正原价
                row['原价'] = real_price
        return row

    # 应用修正（如果有上次价格映射）
    seq = records.iloc[0]['序号']  # 同一组序号相同
    last_price = None
    if last_price_map and seq in last_price_map:
        last_price = last_price_map[seq]
    if last_price is not None:
        records = records.apply(lambda r: fix_price_anomaly(r, last_price), axis=1)

    # 计算基准价格和最终价格（含数量调整）
    records['_base_price'] = records.apply(compute_candidate_price, axis=1)
    adj = records.apply(
        lambda row: adjust_price_by_quantity(row['_base_price'], row.get('关键词'), row.get('货品名称')),
        axis=1, result_type='expand'
    )
    records['_final_price'] = adj[0]
    records['_remark'] = adj[1]

    # 生产日期处理
    records['_parsed_date'] = records['生产日期'].apply(parse_date)
    records['_has_date'] = records['_parsed_date'].notna()

    # 百亿补贴标记
    is_bai = (records['是否百亿补贴产品'].astype(str).str.strip() == '是')
    records['_bai_score'] = is_bai.astype(int)

    # 规格匹配优先级
    if '匹配规格' in records.columns:
        def match_priority(val):
            if pd.isna(val) or val == '':
                return 1
            if '匹配失败' in str(val):
                return 0
            return 2
        records['_spec_priority'] = records['匹配规格'].apply(match_priority)
    else:
        records['_spec_priority'] = 1

    # 排序
    records_sorted = records.sort_values(
        by=['_bai_score', '_has_date', '_parsed_date', '_final_price', '_spec_priority'],
        ascending=[False, False, False, True, False],
        na_position='last'
    )
    return records_sorted.iloc[0]


def apply_excel_style(file_path, highlight_rows=None):
    """
    应用表头样式、居中、自动列宽，并将指定行的价格单元格标浅红色。
    highlight_rows: 需要标红的行索引列表（0-based，不含表头）
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # 找到价格列的列号
        price_col_idx = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == '价格':
                price_col_idx = col
                break

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # 内容居中
        data_align = Alignment(vertical="center", wrap_text=True)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).alignment = data_align

        # 标红
        if highlight_rows and price_col_idx:
            red_fill = PatternFill("solid", fgColor="FFCCCC")
            for row in highlight_rows:
                excel_row = row + 2
                cell = ws.cell(row=excel_row, column=price_col_idx)
                cell.fill = red_fill

        # 自动列宽
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_len = max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
                          default=10)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        wb.save(file_path)
    except Exception as e:
        print(f"应用样式时出错: {e}")


# ================== 主流程 ==================
def main():
    # 读取文件，全部以字符串形式读入
    df_input = pd.read_excel(INPUT_FILE, dtype=str)
    df_collect = pd.read_excel(COLLECT_FILE, dtype=str)

    df_input.columns = df_input.columns.str.strip()
    df_collect.columns = df_collect.columns.str.strip()

    if '序号' not in df_input.columns:
        print("录入表中缺少'序号'列")
        return
    df_input['序号'] = df_input['序号'].astype(str)

    # 确保所需列存在（价格波动复查用）
    if '上次价格' not in df_input.columns:
        df_input['上次价格'] = ''

    last_price_map = {}
    for _, row in df_input.iterrows():
        seq = str(row['序号'])
        last_price = normalize_price(row.get('上次价格'))
        if not pd.isna(last_price):
            last_price_map[seq] = last_price

    required_cols = ['序号', '校验通过', '是否百亿补贴产品', '生产日期', '原价', '现价', '规格价格', '货品名称',
                     '关键词']
    for col in required_cols:
        if col not in df_collect.columns:
            print(f"采集汇总表缺少必要列: {col}")
            return

    # 筛选校验通过的行
    df_collect_valid = df_collect[df_collect['校验通过'] == '✅'].copy()
    if df_collect_valid.empty:
        print("没有校验通过的商品，录入表不会更新")
        df_input.to_excel(OUTPUT_FILE, index=False)
        apply_excel_style(OUTPUT_FILE)
        return

    df_collect_valid['序号'] = df_collect_valid['序号'].astype(str)

    # ---------- 构建保质期映射 ----------
    shelf_life_map = {}
    for _, row in df_input.iterrows():
        seq = str(row['序号'])
        prod_orig = row.get('生产日期', '')
        exp_orig = row.get('失效日期', '')
        days = get_shelf_life_days(prod_orig, exp_orig)
        if days is not None:
            shelf_life_map[seq] = days

    # ---------- 分组选择最佳记录 ----------
    grouped = df_collect_valid.groupby('序号')
    update_map = {}

    for seq, group in grouped:
        best = select_best_record(group, last_price_map)
        if best is None:
            continue

        final_price = best['_final_price']
        remark = best['_remark']

        new_prod_str = best.get('生产日期')
        new_prod_dt = parse_date(new_prod_str)

        new_exp_date = ''
        if pd.notna(new_prod_dt):
            shelf_days = shelf_life_map.get(seq)
            if shelf_days is None:
                shelf_days = 1095
            new_exp_dt = new_prod_dt + timedelta(days=shelf_days)
            new_exp_date = new_exp_dt.strftime('%Y-%m-%d')
        else:
            new_exp_date = ''

        formatted_prod = new_prod_dt.strftime('%Y-%m-%d') if pd.notna(new_prod_dt) else ''
        is_bai = best.get('是否百亿补贴产品', '')

        if isinstance(final_price, (list, np.ndarray, pd.Series)):
            final_price = final_price.iloc[0] if hasattr(final_price, 'iloc') else float(final_price[0])

        update_map[seq] = {
            '价格': float(final_price) if not pd.isna(final_price) else np.nan,
            '生产日期': formatted_prod,
            '是否百亿补贴产品': is_bai,
            '失效日期': new_exp_date,
            '备注': remark
        }

    # ---------- 更新录入表 ----------
    for col in ['价格', '生产日期', '是否百亿补贴产品', '失效日期', '备注']:
        if col not in df_input.columns:
            df_input[col] = ''
    for col in ['价格', '生产日期', '是否百亿补贴产品', '失效日期', '备注']:
        if col in df_input.columns:
            df_input[col] = df_input[col].astype(object)

    for idx, row in df_input.iterrows():
        seq = str(row['序号'])
        if seq in update_map:
            updates = update_map[seq]
            df_input.at[idx, '价格'] = updates['价格']
            df_input.at[idx, '生产日期'] = updates['生产日期']
            df_input.at[idx, '是否百亿补贴产品'] = updates['是否百亿补贴产品']
            df_input.at[idx, '失效日期'] = updates['失效日期']
            df_input.at[idx, '备注'] = updates['备注']

    # 价格列转为数值
    df_input['价格'] = pd.to_numeric(df_input['价格'], errors='coerce')
    df_input['上次价格'] = pd.to_numeric(df_input['上次价格'], errors='coerce')

    # ---------- 价格波动人工复查（仅与上次价格比较） ----------
    # 价格波动人工复查（仅与上次价格比较，200以下且波动绝对值<=70豁免）
    highlight_indices = []
    for idx, row in df_input.iterrows():
        price = row['价格']
        if pd.isna(price):
            continue
        last_price = row['上次价格']
        if pd.isna(last_price):
            continue
        # 判断是否超过比例阈值
        if price > last_price * 1.5 or price < last_price * (2 / 3):
            # 价格<200 且 波动绝对值<=70 → 豁免不标红
            if price < 200 and abs(price - last_price) <= 70:
                continue
            highlight_indices.append(idx)

    # 价格保留两位小数
    df_input['价格'] = df_input['价格'].round(2)

    # 输出 Excel
    df_input.to_excel(OUTPUT_FILE, index=False)
    apply_excel_style(OUTPUT_FILE, highlight_rows=highlight_indices)
    print(f"处理完成，结果已保存至: {OUTPUT_FILE}")
    if highlight_indices:
        print(f"注意：以下行号（1-based，含表头）的价格单元格已标浅红色：{[i + 2 for i in highlight_indices]}")


if __name__ == "__main__":
    main()