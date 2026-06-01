import pandas as pd
import numpy as np
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ================== 配置 ==================
INPUT_FILE = "录入表.xlsx"  # 原始录入表
COLLECT_FILE = "商品采集汇总.xlsx"  # 采集汇总表
OUTPUT_FILE = "录入表_更新.xlsx"  # 输出文件

# 需要检测的数量词（双份）
QUANTITY_PATTERN = re.compile(r'(双支|两支|2支|\*2|对装|双只|两瓶|两支装|双支装)', re.IGNORECASE)


# ================== 辅助函数 ==================
def normalize_price(value):
    """将价格列转换为浮点数，处理空值、'-'等"""
    if pd.isna(value):
        return np.nan
    if isinstance(value, str):
        value = value.strip().replace('¥', '').replace('元', '')
        if value == '' or value == '-':
            return np.nan
    try:
        return float(value)
    except:
        return np.nan


def parse_date(date_val):
    """将各种格式的日期转换为 pandas datetime，无效则返回 NaT"""
    if pd.isna(date_val):
        return pd.NaT
    try:
        # 处理字符串格式如 "2025/6/1" 或 "2024-07-01"
        return pd.to_datetime(date_val, errors='coerce')
    except:
        return pd.NaT


def adjust_price_by_quantity(price, keyword, product_name):
    """
    根据关键词和货品名称中的数量词调整价格。
    规则：关键词无数量词 but 货品名称有 -> 除以2
          关键词有数量词 but 货品名称无 -> 乘以2
          其他情况不变
    """
    if pd.isna(price):
        return price
    kw_has = bool(QUANTITY_PATTERN.search(str(keyword))) if not pd.isna(keyword) else False
    name_has = bool(QUANTITY_PATTERN.search(str(product_name))) if not pd.isna(product_name) else False

    if not kw_has and name_has:
        return price / 2.0
    elif kw_has and not name_has:
        return price * 2.0
    else:
        return price


def compute_candidate_price(row):
    """
    根据采集汇总表中的一行数据，计算出用于比较和最终录入的价格（尚未做数量调整）
    优先级：规格价格 > (百亿补贴?原价:现价)
    """
    # 规格价格最高优先级
    spec_price = normalize_price(row.get('规格价格', np.nan))
    if not pd.isna(spec_price):
        return spec_price

    # 没有规格价格：根据是否百亿补贴决定取原价还是现价
    is_bai = str(row.get('是否百亿补贴产品', '')).strip() == '是'
    orig_price = normalize_price(row.get('原价', np.nan))
    curr_price = normalize_price(row.get('现价', np.nan))

    if is_bai:
        # 百亿补贴：优先原价，原价缺失则用现价
        return orig_price if not pd.isna(orig_price) else curr_price
    else:
        # 非百亿补贴：优先现价
        return curr_price if not pd.isna(curr_price) else orig_price


def select_best_record(records):
    """
    从同一序号的多条记录中选出最佳一条。
    records: DataFrame 子集，已通过校验通过筛选。
    返回最佳记录的 Series 索引，若无可选记录返回 None。
    """
    if records.empty:
        return None

    # 为每条记录计算基准价格（未做数量调整）和最终价格（调整后）
    records = records.copy()
    # 基准价格：规格价 > (百亿补贴?原价:现价)
    records['_base_price'] = records.apply(compute_candidate_price, axis=1)
    # 最终价格：基准价格经过数量调整（乘/除2）
    records['_final_price'] = records.apply(
        lambda row: adjust_price_by_quantity(row['_base_price'], row.get('关键词'), row.get('货品名称')),
        axis=1
    )
    # 生产日期处理
    records['_parsed_date'] = records['生产日期'].apply(parse_date)
    records['_has_date'] = records['_parsed_date'].notna()
    # 百亿补贴得分
    records['_bai_score'] = (records['是否百亿补贴产品'].astype(str).str.strip() == '是').astype(int)

    # 按优先级排序：百亿补贴 > 有生产日期 > 生产日期越晚越好 > 价格越低越好
    records_sorted = records.sort_values(
        by=['_bai_score', '_has_date', '_parsed_date', '_final_price'],
        ascending=[False, False, False, True],
        na_position='last'
    )
    best = records_sorted.iloc[0]
    return best


# ================== 主流程 ==================
def main():
    # 1. 读取 Excel
    try:
        df_input = pd.read_excel(INPUT_FILE, dtype=str)  # 先全部读为字符串，避免类型混乱
        df_collect = pd.read_excel(COLLECT_FILE, dtype=str)
    except Exception as e:
        print(f"读取文件失败: {e}")
        return

    # 清理列名（去除首尾空格）
    df_input.columns = df_input.columns.str.strip()
    df_collect.columns = df_collect.columns.str.strip()

    # 确保录入表有序号列，且为字符串方便匹配
    if '序号' not in df_input.columns:
        print("录入表中缺少'序号'列")
        return
    df_input['序号'] = df_input['序号'].astype(str)

    # 采集表必须的列
    required_cols = ['序号', '校验通过', '是否百亿补贴产品', '生产日期', '原价', '现价', '规格价格', '货品名称',
                     '关键词']
    for col in required_cols:
        if col not in df_collect.columns:
            print(f"采集汇总表缺少必要列: {col}")
            return

    # 2. 筛选校验通过的行
    df_collect_valid = df_collect[df_collect['校验通过'] == '✅'].copy()
    if df_collect_valid.empty:
        print("没有校验通过的商品，录入表不会更新")
        # 仍然保存原表（或输出空更新表）
        df_input.to_excel(OUTPUT_FILE, index=False)
        apply_excel_style(OUTPUT_FILE)
        return

    # 3. 按序号分组，为每个序号选择最佳记录
    # 将序号统一为字符串
    df_collect_valid['序号'] = df_collect_valid['序号'].astype(str)
    grouped = df_collect_valid.groupby('序号')

    # 记录更新映射：序号 -> {价格, 生产日期, 是否百亿补贴产品}
    update_map = {}
    for seq, group in grouped:
        best = select_best_record(group)
        if best is not None:
            # 计算最终价格（已含数量调整）
            raw_price = compute_candidate_price(best)
            final_price = adjust_price_by_quantity(raw_price, best.get('关键词'), best.get('货品名称'))
            # 生产日期格式转换：2024-05-01 -> 2024/05/01
            prod_date = best.get('生产日期')
            if pd.notna(prod_date) and str(prod_date).strip():
                # 尝试解析并重新格式化
                dt = parse_date(prod_date)
                if pd.notna(dt):
                    prod_date = dt.strftime('%Y/%m/%d')
                else:
                    # 如果解析失败，直接替换 - 为 /
                    prod_date = str(prod_date).replace('-', '/')
            else:
                prod_date = ''
            # 是否百亿补贴
            is_bai = best.get('是否百亿补贴产品', '')
            update_map[seq] = {
                '价格': final_price,
                '生产日期': prod_date,
                '是否百亿补贴产品': is_bai
            }

    # 4. 更新录入表
    # 转换价格列为浮点数以便后续写入
    if '价格' in df_input.columns:
        df_input['价格'] = df_input['价格'].apply(normalize_price)
    else:
        df_input['价格'] = np.nan

    # 对每一行应用更新
    for idx, row in df_input.iterrows():
        seq = str(row['序号'])
        if seq in update_map:
            updates = update_map[seq]
            df_input.at[idx, '价格'] = updates['价格']
            df_input.at[idx, '生产日期'] = updates['生产日期']
            df_input.at[idx, '是否百亿补贴产品'] = updates['是否百亿补贴产品']

    # 5. 输出到 Excel，并应用样式
    # 注意：浮点数列保留两位小数（可选）
    if '价格' in df_input.columns:
        df_input['价格'] = df_input['价格'].round(2)

    df_input.to_excel(OUTPUT_FILE, index=False)
    apply_excel_style(OUTPUT_FILE)
    print(f"处理完成，结果已保存至: {OUTPUT_FILE}")


def apply_excel_style(file_path):
    """应用表头样式、居中、自动列宽"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # 内容居中
        data_align = Alignment(vertical="center", wrap_text=True)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).alignment = data_align

        # 自动列宽
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_len = max(
                (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
                default=10
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        wb.save(file_path)
    except Exception as e:
        print(f"应用样式时出错: {e}")


if __name__ == "__main__":
    main()