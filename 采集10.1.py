import uiautomator2 as u2
import cv2
from ultralytics import YOLO
import time
import re
import requests
import easyocr
import pandas as pd
import os
import logging
import subprocess
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from product_validator import validate_product   # 导入新版校验
import sku_matcher   # 新增导入
import traceback

logging.disable(logging.WARNING)
YOLO().verbose = False

# ====================== 初始化 ======================
d = u2.connect()
model = YOLO("runs/detect/pdd_logo_train-2/weights/best.pt")
subsidy_model = YOLO("runs/detect/subsidy_train/weights/best.pt")
detail_model = YOLO("runs/detect/product_detail_train/weights/best.pt")
reader = easyocr.Reader(['ch_sim'], gpu=False)

# ====================== 配置项 ======================
PRODUCT_LIST_FILE = "模块开发/测试用例.xlsx"
SEARCH_INTERVAL_SECONDS = 60
PACKAGE_NAME = "com.xunmeng.pinduoduo"

# ====================== 全局存储 ======================
record_list = []
debug_record_list = []

EXCEL_HEADER = [
    "序号", "货品名称", "关键词", "原价", "现价",
    "是否百亿补贴产品", "生产日期", "校验通过", "未通过原因",
    "匹配规格", "规格价格"          # 新增两列
]

DEBUG_EXCEL_HEADER = [
    "校验序号", "搜索关键词", "提取到的商品标题",
    "搜索词匹配品牌", "商品标题匹配品牌",
    "品牌是否一致", "规格是否匹配通过",
    "品名匹配率(%)", "最终校验是否通过",
    "校验时间", "备注/失败原因",
    "匹配规格", "规格价格"
]

# ====================== 保存调试记录 ======================
def save_debug_excel():
    if not debug_record_list:
        return
    file_path = "商品校验调试记录.xlsx"
    new_df = pd.DataFrame(debug_record_list, columns=DEBUG_EXCEL_HEADER)
    if os.path.exists(file_path):
        try:
            old_df = pd.read_excel(file_path)
            new_df = pd.concat([old_df, new_df], ignore_index=True)
        except:
            pass
    new_df.drop_duplicates(subset=["搜索关键词", "提取到的商品标题", "校验时间"], keep="last", inplace=True)
    new_df["校验序号"] = list(range(1, len(new_df) + 1))
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        new_df.to_excel(writer, index=False, sheet_name='校验记录')
        workbook = writer.book
        worksheet = writer.sheets['校验记录']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        data_align = Alignment(vertical="center", wrap_text=True)
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = data_align
                if col == 8:
                    cell.number_format = '0.00"%"'
        for col in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    current_length = len(str(cell.value))
                    if current_length > max_length:
                        max_length = current_length
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
    print(f"📋 校验记录已保存，总记录数：{len(new_df)}")

def save_all_to_excel():
    if record_list:
        df = pd.DataFrame(record_list, columns=EXCEL_HEADER)
        file = "商品采集汇总.xlsx"
        if os.path.exists(file):
            old_df = pd.read_excel(file)
            df = pd.concat([old_df, df], ignore_index=True)
        df = df.drop_duplicates(subset=["序号", "货品名称", "现价"], keep="last")

        # 带样式写入
        with pd.ExcelWriter(file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="汇总")
            ws = writer.sheets["汇总"]


            # 表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4472C4")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align

            # 内容居中
            data_align = Alignment(vertical="center", wrap_text=True)
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).alignment = data_align

            # 自动列宽
            for col in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col)
                max_len = max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)), default=10)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        print(f"📁 采集记录已保存，总行数：{len(df)}")
    save_debug_excel()

# ====================== 辅助函数（未修改） ======================
def go_to_pinduoduo_home():
    print("\n🔴 强制返回首页...")
    try:
        subprocess.run(["adb", "shell", "am", "start", "-S", "-n",
                        "com.xunmeng.pinduoduo/com.xunmeng.pinduoduo.ui.activity.MainFrameActivity"],
                       check=True, timeout=15)
        time.sleep(6)
        print("🟢 已回到首页")
    except:
        subprocess.run(["adb", "shell", "am", "force-stop", PACKAGE_NAME])
        time.sleep(2)
        subprocess.run(["adb", "shell", "am", "start", "-n",
                        "com.xunmeng.pinduoduo/com.xunmeng.pinduoduo.ui.activity.MainFrameActivity"])
        time.sleep(8)

def load_product_list_with_status():
    if not os.path.exists(PRODUCT_LIST_FILE):
        print("❌ 名单文件不存在")
        return []
    df = pd.read_excel(PRODUCT_LIST_FILE)
    if "状态" not in df.columns:
        df["状态"] = "未采集"
    if "序号" not in df.columns:
        df["序号"] = range(1, len(df)+1)
        df.to_excel(PRODUCT_LIST_FILE, index=False)
    todo = df[df["状态"] == "未采集"].copy()
    print(f"✅ 总商品数：{len(df)} | 待采集：{len(todo)}")
    return todo

def mark_product_as_done_by_index(index_num):
    try:
        df = pd.read_excel(PRODUCT_LIST_FILE)
        df.loc[df["序号"] == index_num, "状态"] = "已采集"
        df.to_excel(PRODUCT_LIST_FILE, index=False)
        print(f"✅ 序号 {index_num} 已标记为已采集")
    except Exception as e:
        print(f"❌ 标记失败：{e}")
        backup_name = f"{PRODUCT_LIST_FILE}.bak"
        os.rename(PRODUCT_LIST_FILE, backup_name)
        print(f"⚠️ 文件已备份为 {backup_name}")

# ====================== 搜索与列表识别（保留原逻辑） ======================
def search_product(keyword):
    print(f"\n🔍 搜索：{keyword}")
    width, height = d.window_size()
    search_y = int(height * 200 / 2400)
    d.drag(width // 2, height // 2, width // 2, height // 2 + 400)
    time.sleep(1)
    d.click(width // 2, search_y)
    time.sleep(0.8)
    d(className="android.widget.EditText").clear_text()
    d(className="android.widget.EditText").set_text(keyword)
    time.sleep(0.8)
    d(text="搜索", className="android.widget.TextView").click()
    time.sleep(3)

def scan_list_products():
    d.screenshot("list_screen.jpg")
    img = cv2.imread("list_screen.jpg")
    results = model(img, conf=0.2)
    products = []
    for r in results:
        for box in r.boxes:
            cls = int(box.cls)
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
            if cls == 0:
                products.append({"type": "baiyi", "cx": cx, "cy": cy})
            elif cls == 1:
                products.append({"type": "brand", "cx": cx, "cy": cy})
            elif cls == 2:
                products.append({"type": "global", "cx": cx, "cy": cy})
    return products

def get_products_with_tags():
    prods = scan_list_products()
    grouped = {}
    for p in prods:
        cx, cy = p["cx"], p["cy"]
        matched = False
        for key in list(grouped.keys()):
            ecx, ecy = key
            if abs(cx - ecx) < 200 and abs(cy - ecy) < 180:
                grouped[key]["tags"].add(p["type"])
                matched = True
                break
        if not matched:
            grouped[(cx, cy)] = {"tags": {p["type"]}, "cx": cx, "cy": cy}
    return list(grouped.values())

def get_priority(tags):
    """单个百亿补贴与百亿补贴+品牌同等最高优先级"""
    if "baiyi" in tags:          # 包含百亿补贴就是最高
        return 4
    if "brand" in tags:
        return 2
    if "global" in tags:
        return 1
    return 0

def is_all_global(item_list):
    return all(not ("baiyi" in i["tags"] or "brand" in i["tags"]) for i in item_list)

def scroll_down_once():
    width, height = d.window_size()
    # 自适应屏幕 向上滑动（从屏幕 85% 位置 拖到 25% 位置）
    d.drag(
        width * 0.5,  # 起点 X：屏幕中间
        height * 0.85,  # 起点 Y：下方 85%
        width * 0.5,  # 终点 X：屏幕中间
        height * 0.25,  # 终点 Y：上方 25%
        duration=0.3
    )
    time.sleep(2.5)

def scroll_to_top():
    width, height = d.window_size()
    for _ in range(2):
        d.drag(width // 2, height * 0.2, width // 2, height * 0.8, 0.3)
        time.sleep(0.8)

def sort_products_by_priority():
    raw = get_products_with_tags()
    if is_all_global(raw):
        scroll_down_once()
        raw = get_products_with_tags()
        if is_all_global(raw):
            scroll_down_once()
            raw = get_products_with_tags()
            if is_all_global(raw):
                scroll_to_top()
                raw = get_products_with_tags()
    s = sorted(raw, key=lambda x: get_priority(x["tags"]), reverse=True)
    if is_all_global(s):
        s = s[:2]
    return s

def is_subsidy_product():
    return "百亿补贴" in d.dump_hierarchy() or "官方补贴" in d.dump_hierarchy()

def extract_product_info(xml_content: str, search_word: str):
    def get_ngram_pairs(text, n=2):
        text = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", text.lower())
        return (
            [text[i : i + n] for i in range(len(text) - n + 1)]
            if len(text) >= n
            else [text]
        )

    def get_single_chars(text):
        text = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", text.lower())
        return [c for c in text]

    def count_chinese(text):
        return len(re.findall(r"[\u4e00-\u9fff]", text))

    search_cn_count = count_chinese(search_word)
    desc_list = re.findall(r'content-desc="([^"]+)"', xml_content)
    best_title = ""
    best_count = 0
    blacklist = [
        "电池", "状态栏", "电量", "百分之", "WLAN", "信号",
        "通知", "高德", "淘宝", "浏览器", "手机管家", "振铃器", "静音",
        "返回", "分享", "店铺", "客服", "工具栏", "顶部", "拼小圈",
        "¥", "￥", "大促价", "已抢", "假一赔十", "100%正品", "拼单价",
        "狂降", "直接成团", "买过", "次", "图片", "该店", "tronplayer_view", "查看全部",
    ]
    search_pairs = get_ngram_pairs(search_word)
    for desc in desc_list:
        desc = desc.strip()
        if any(kw in desc for kw in blacklist):
            continue
        desc_clean = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", desc.lower())
        match_count = sum(1 for p in search_pairs if p in desc_clean)
        if match_count > best_count and match_count > 0:
            best_count = match_count
            best_title = desc
        elif match_count == best_count and match_count > 0:
            if len(desc) > len(best_title):
                best_title = desc
    if not best_title:
        search_chars = get_single_chars(search_word)
        best_count = 0
        for desc in desc_list:
            desc = desc.strip()
            if any(kw in desc for kw in blacklist):
                continue
            if count_chinese(desc) < search_cn_count:
                continue
            desc_clean = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", desc.lower())
            match_count = sum(1 for c in search_chars if c in desc_clean)
            if match_count > best_count and match_count > 0:
                best_count = match_count
                best_title = desc
            elif match_count == best_count and match_count > 0:
                if len(desc) > len(best_title):
                    best_title = desc

    # ====================== 以下是【全新重写】的价格提取逻辑 ======================
    # 1. 找到所有带 ¥ 的 content-desc 文本
    price_desc_list = []
    for desc in re.findall(r'content-desc="([^"]+)"', xml_content):
        if "¥" in desc:
            price_desc_list.append(desc)

    # 2. 清洗：新增2条规则 + 原有规则
    cleaned_texts = []
    for text in price_desc_list:
        if "单独购买" in text:
            continue
        # ====================== 清洗规则（截断版） ======================
        cleaned_texts = []
        for text in price_desc_list:
            # 规则0：包含单独购买 → 跳过
            if "单独购买" in text:
                continue
            # 找到所有规则的最早出现位置
            split_at = len(text)
            # 规则1：时间 xx:xx → 截断
            match1 = re.search(r"\d{1,2}:\d{2}", text)
            if match1:
                split_at = min(split_at, match1.start())
            # 规则2：X人团 → 截断
            match2 = re.search(r"\d人团", text)
            if match2:
                split_at = min(split_at, match2.start())
            # 规则3：数字+元 / 件 / 万 → 截断
            match3 = re.search(r"\d+\.?\d*(元|件|万\+?|万)", text)
            if match3:
                split_at = min(split_at, match3.start())
            # 规则4：降 + 数字 → 截断
            match4 = re.search(r"降\d+\.?\d*", text)
            if match4:
                split_at = min(split_at, match4.start())

            match5 = re.search(r"\d{2}人想拼", text)
            if match5:
                split_at = min(split_at, match5.start())
            # 从最早匹配的位置截断，后面全部丢掉
            text = text[:split_at].strip()
            cleaned_texts.append(text)

    # 3. 统一提取所有价格
    all_prices = []
    for t in cleaned_texts:
        prices = re.findall(r"[¥￥]\s*(\d+\.?\d*)", t)
        all_prices.extend(prices)

    # 4. 过滤规则：去掉 0开头 / 个位数（1-9）
    valid_prices = []
    for p in all_prices:
        p_str = str(p).strip()
        # 跳过空
        if not p_str:
            continue
        # 跳过 0 开头
        if p_str.startswith("0") and len(p_str) > 1:
            continue
        # 转数字
        try:
            num = float(p_str)
        except:
            continue
        # 跳过个位数
        if num < 10:
            continue
        valid_prices.append(round(num, 2))

    # 去重 + 排序
    valid_prices = sorted(list(set(valid_prices)))

    original_price = None
    current_price = None
    if len(valid_prices) >= 2:
        original_price = str(max(valid_prices))
        current_price = str(min(valid_prices))
    elif len(valid_prices) == 1:
        current_price = str(valid_prices[0])  # 只有一个 → 算现价

    return {
        "title": best_title.strip() if best_title else "",
        "original_price": original_price,
        "current_price": current_price,
    }

def find_and_click_detail(max_scroll=7):
    for _ in range(max_scroll):
        img = d.screenshot(format="opencv")
        res = detail_model(img, conf=0.75)
        box = None
        for r in res:
            for b in r.boxes:
                if int(b.cls[0]) == 0:
                    box = tuple(map(int, b.xyxy[0]))
                    break
            if box: break
        if box:
            x1, y1, x2, y2 = box
            crop_img = img[y1:y2, x1:x2]
            ocr_result = reader.readtext(crop_img)
            full_text = "".join(item[1] for item in ocr_result)
            if "商品详情" in full_text:
                cx = (x1 + x2) // 2
                cy = (y1 + y2) // 2
                d.click(cx, cy)
                time.sleep(1.5)
                if "生产日期" in d.dump_hierarchy():
                    return True
                else:
                    d.press("back")
                    return False
        width, height = d.window_size()
        # 原固定坐标：(500,1800)→(500,600) 向上拖
        d.drag(
            int(width * 0.46),
            int(height * 0.75),
            int(width * 0.46),
            int(height * 0.25),
            0.25
        )
        time.sleep(0.8)
    return False

def get_date_with_retry():
    m = re.search(r'text="(\d{4}-\d{1,2}-\d{1,2})"', d.dump_hierarchy())
    return m.group(1) if m else ""

def collect_single_product(search_word, serial_num):
    xml = d.dump_hierarchy()
    info = extract_product_info(xml, search_word)
    subsidy = "是" if is_subsidy_product() else "否"

    title = info["title"]
    ori = info["original_price"]
    cur = info["current_price"]

    # 执行商品信息校验
    res = validate_product(search_word, title)
    match_pass = res['final']
    fail_reason = "" if match_pass else res.get("remark", "校验未通过")

    matched_spec = ""
    spec_price = ""
    _, title_color_codes = sku_matcher.extract_specs(title)  # 注意 extract_specs 返回 (cap_nums, color_codes)

    # 触发条件：校验失败且原因含规格 或 标题色号数量大于1
    if (not match_pass and any(kw in fail_reason for kw in ["规格", "色号", "容量", "浓度"])) or len(title_color_codes) > 1:
        if len(title_color_codes) > 1:
            print(f"🔧 标题包含多规格 {title_color_codes}，尝试进入规格面板匹配...")
        else:
            print(f"🔧 校验失败原因为规格不匹配，尝试进入规格面板匹配...")
        try:
            width, height = d.window_size()
            d.click(int(width * 0.8), int(height * 0.96))
            time.sleep(1.5)
            spec_result = sku_matcher.get_sku_price_auto(d, search_word, click_timeout=1.5)
            if spec_result.get("current_price"):
                matched_spec = spec_result["title"]
                spec_price = spec_result["current_price"]
                cur = spec_price
                # 关键修改：只有原品名匹配通过时，才认为校验通过
                if res.get('name_ok', False):
                    match_pass = True
                    fail_reason = ""
                    print(f"✅ 规格匹配成功，且品名已匹配，校验通过")
                else:
                    match_pass = False
                    # 保留原失败原因，或追加说明
                    fail_reason = res.get("remark", "品名不匹配，规格匹配但无效")
                    print(f"⚠️ 规格匹配成功但品名不匹配，校验仍不通过")
            else:
                print("⚠️ 规格匹配未获取到有效价格，维持原校验结果")
            d.press("back")
            time.sleep(1)
        except Exception as e:
            print(f"❌ 规格匹配异常：{e}")
            d.press("back")
            time.sleep(1)

    # 获取生产日期
    detail = find_and_click_detail()
    date = get_date_with_retry() if detail else ""

    # 写入记录
    record_list.append([
        serial_num, title, search_word, ori, cur, subsidy, date,
        "✅" if match_pass else "❌",
        fail_reason,
        matched_spec,
        spec_price
    ])

    debug_record_list.append([
        len(debug_record_list) + 1,
        search_word,
        title,
        res['s_brand'],
        res['p_brand'],
        "是" if res['brand_ok'] else "否",
        "是" if res['spec_ok'] else "否",
        round(res['ratio'] * 100, 2),
        "是" if match_pass else "否",  # 使用匹配后的结果
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        fail_reason,
        matched_spec,
        spec_price
    ])

    print("\n" + "=" * 80)
    print(f"货品名称：{title}")
    print(f"关键词：{search_word}")
    print(f"原价：{ori} | 现价：{cur}")
    if matched_spec:
        print(f"匹配规格：{matched_spec} | 规格价格：{spec_price}")
    print(f"百亿补贴：{subsidy} | 日期：{date}")
    print(f"校验通过：{'✅' if match_pass else '❌'} {match_pass}")
    print("=" * 80)

    return {
        "title": title,
        "subsidy": subsidy,
        "date": date,
        "found_detail": detail,
        "passed": match_pass,
        "matched_spec": matched_spec,
        "spec_price": spec_price
    }

def ensure_back_to_list(d):
    """确保返回到列表页，处理可能的优惠券弹窗"""
    time.sleep(0.5)
    try:
        if d(textContains="放弃优惠").exists or d(textContains="继续退出").exists:
            print("⚠️ 检测到优惠券弹窗，再次返回")
            d.press("back")
            time.sleep(0.5)
    except Exception as e:
        print(f"ensure_back_to_list 异常: {e}")

def select_and_collect_best_product(search_word, serial_num):
    sorted_prods = sort_products_by_priority()
    if not sorted_prods:
        print("❌ 未识别商品")
        return None

    # 按优先级分组，从高到低依次尝试
    priority_order = sorted(set(get_priority(p["tags"]) for p in sorted_prods), reverse=True)
    any_passed = False

    for prio in priority_order:
        candidates = [p for p in sorted_prods if get_priority(p["tags"]) == prio]
        if not candidates:
            continue
        # ====================== 全球购特殊处理：回到顶部 + 重新识别（不触发自动下滑）
        if prio == 1:
            print("🔽 降级到全球购，返回顶部并重新识别（不自动下滑）")
            scroll_to_top()
            time.sleep(1)
            # 直接获取顶部商品，不调用 sort_products_by_priority（避免下滑）
            raw = get_products_with_tags()
            sorted_prods = sorted(raw, key=lambda x: get_priority(x["tags"]), reverse=True)
            candidates = [p for p in sorted_prods if get_priority(p["tags"]) == prio]
            candidates = candidates[:3]  # 最多采3个
        print(f"\n===== 优先级 {prio}，共 {len(candidates)} 个商品 =====")
        time.sleep(0.5)

        # 采集该优先级所有商品
        for i, p in enumerate(candidates):
            print(f"--- 进入商品 {i + 1}/{len(candidates)} ---")
            d.click(p["cx"], p["cy"])
            time.sleep(0.5)
            res = collect_single_product(search_word, serial_num)
            if res["passed"]:
                any_passed = True
            # 返回列表页（处理弹窗）
            if res["found_detail"]:
                d.press("back")
                ensure_back_to_list(d)
                time.sleep(1.5)
            d.press("back")
            ensure_back_to_list(d)
            time.sleep(1)

        # 如果当前优先级已有商品通过，不再降级（但仍会采集完当前级所有商品）
        if any_passed:
            print(f"✅ 优先级 {prio} 中已有商品通过校验，停止降级")
            break
        else:
            print(f"⚠️ 优先级 {prio} 全部未通过，尝试下一优先级")

    return any_passed

# ====================== 进度条工具 ======================
def wait_with_progress(seconds):
    import sys
    bar_len = 30
    for i in range(seconds):
        filled = bar_len * (i + 1) // seconds
        bar = "█" * filled + "-" * (bar_len - filled)
        percent = (i + 1) / seconds * 100
        sys.stdout.write(f"\r⏳ 等待下一个商品：[{bar}] {percent:.0f}% ({i+1}/{seconds}s)")
        sys.stdout.flush()
        time.sleep(1)
    sys.stdout.write("\n")

# ====================== 主循环 ======================
def main():
    print("🚀 启动采集 + 品牌规格品名四重校验")
    todo = load_product_list_with_status()
    if todo.empty:
        print("🎉 全部完成")
        return
    for _, row in todo.iterrows():
        kw = str(row["货品名称"]).strip()
        idx = int(row["序号"])
        ok = False
        while not ok:
            try:
                search_product(kw)
                select_and_collect_best_product(kw, idx)
                save_all_to_excel()
                mark_product_as_done_by_index(idx)
                # ========== 这里换成带进度条的等待 ==========
                print(f"\n⏸ 等待 {SEARCH_INTERVAL_SECONDS} 秒后继续...")
                wait_with_progress(SEARCH_INTERVAL_SECONDS)
                ok = True
            except Exception as e:
                print(f"❌ 异常：{e}")
                # 获取报错行号
                traceback.print_exc()
                go_to_pinduoduo_home()
    print("🎉 全部采集完成！")

if __name__ == "__main__":
    main()