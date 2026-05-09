import logging
import os
import re
import subprocess
import time
import traceback
from datetime import datetime
from typing import Any, Dict, List, Optional

import argparse
import json
import cv2
import easyocr
import pandas as pd
import threading
import uiautomator2 as u2
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from ultralytics import YOLO

from product_validator import validate_product

logging.disable(logging.WARNING)

_PROJECT_DIR = os.path.abspath(
    os.environ.get("COLLECT_PROJECT_DIR") or os.path.dirname(os.path.abspath(__file__))
)
PRODUCT_LIST_FILE = os.path.abspath(
    os.environ.get("COLLECT_TASK_XLSX") or os.path.join(_PROJECT_DIR, "搜索名单.xlsx")
)
SUMMARY_XLSX = os.path.abspath(
    os.environ.get("COLLECT_SUMMARY_XLSX") or os.path.join(_PROJECT_DIR, "商品采集汇总.xlsx")
)
DEBUG_XLSX = os.path.abspath(
    os.environ.get("COLLECT_DEBUG_XLSX") or os.path.join(_PROJECT_DIR, "商品校验调试记录.xlsx")
)

SEARCH_INTERVAL_SECONDS = int(os.environ.get("COLLECT_SEARCH_INTERVAL", "40"))
PACKAGE_NAME = "com.xunmeng.pinduoduo"

_runtime_lock = threading.Lock()
_infer_lock = threading.Lock()
_runtime: Optional[Dict[str, Any]] = None

record_list: List[Any] = []
debug_record_list: List[Any] = []
_OUTPUT_LOCK = threading.RLock()

EXCEL_HEADER = [
    "序号",
    "货品名称",
    "关键词",
    "原价",
    "现价",
    "是否百亿补贴产品",
    "生产日期",
]

DEBUG_EXCEL_HEADER = [
    "校验序号",
    "搜索关键词",
    "提取到的商品标题",
    "搜索词匹配品牌",
    "商品标题匹配品牌",
    "品牌是否一致",
    "规格是否匹配通过",
    "品名匹配率(%)",
    "最终校验是否通过",
    "校验时间",
    "备注/失败原因",
]


def _ensure_project_cwd() -> None:
    try:
        os.chdir(_PROJECT_DIR)
    except Exception:
        pass


def _get_runtime() -> Dict[str, Any]:
    global _runtime
    with _runtime_lock:
        if _runtime is None:
            _ensure_project_cwd()
            logo = YOLO("runs/detect/pdd_logo_train-2/weights/best.pt")
            detail = YOLO("runs/detect/product_detail_train/weights/best.pt")
            for m in (logo, detail):
                try:
                    m.verbose = False
                except Exception:
                    pass
            _runtime = {
                "model": logo,
                "detail_model": detail,
                "reader": easyocr.Reader(["ch_sim"], gpu=False),
            }
        return _runtime


def _emit(
    signal: Optional[Dict[str, Any]],
    phase: str,
    detail: str = "",
    extra: Optional[Dict[str, Any]] = None,
) -> None:
    if signal is None:
        return
    signal["阶段"] = phase
    signal["阶段说明"] = detail
    signal["更新时间"] = time.strftime("%Y-%m-%d %H:%M:%S")
    if extra:
        signal.update(extra)


def _write_debug_to_disk() -> None:
    if not debug_record_list:
        return
    file_path = DEBUG_XLSX
    new_df = pd.DataFrame(debug_record_list, columns=DEBUG_EXCEL_HEADER)
    if os.path.exists(file_path):
        try:
            old_df = pd.read_excel(file_path)
            new_df = pd.concat([old_df, new_df], ignore_index=True)
        except Exception:
            pass
    new_df.drop_duplicates(
        subset=["搜索关键词", "提取到的商品标题", "校验时间"],
        keep="last",
        inplace=True,
    )
    new_df["校验序号"] = list(range(1, len(new_df) + 1))
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        new_df.to_excel(writer, index=False, sheet_name="校验记录")
        worksheet = writer.sheets["校验记录"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        data_align = Alignment(vertical="center", wrap_text=True)
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = data_align
                if col == 8:
                    cell.number_format = '0.00"%"'
        for col in range(1, worksheet.max_column + 1):
            col_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    current_length = len(str(cell.value))
                    if current_length > max_length:
                        max_length = current_length
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
    print(f"校验记录已保存，总记录数：{len(new_df)}")


def _write_summary_to_disk() -> None:
    if not record_list:
        return
    df = pd.DataFrame(record_list, columns=EXCEL_HEADER)
    file = SUMMARY_XLSX
    if os.path.exists(file):
        old_df = pd.read_excel(file)
        df = pd.concat([old_df, df], ignore_index=True)
    df = df.drop_duplicates(subset=["序号", "货品名称", "现价"], keep="last")
    df.to_excel(file, index=False)
    print(f"采集记录已保存，总行数：{len(df)}")


def save_debug_excel() -> None:
    with _OUTPUT_LOCK:
        _write_debug_to_disk()


def save_all_to_excel() -> None:
    with _OUTPUT_LOCK:
        _write_summary_to_disk()
        _write_debug_to_disk()


class DeviceContext:
    def __init__(self, device_id: str):
        self.device_id = device_id
        self.d = u2.connect(device_id)


def go_to_pinduoduo_home(device_id: str) -> None:
    print("\n强制返回首页...")
    try:
        subprocess.run(
            [
                "adb",
                "-s",
                device_id,
                "shell",
                "am",
                "start",
                "-S",
                "-n",
                "com.xunmeng.pinduoduo/com.xunmeng.pinduoduo.ui.activity.MainFrameActivity",
            ],
            check=True,
            timeout=15,
        )
        time.sleep(6)
        print("已回到首页")
    except Exception:
        subprocess.run(["adb", "-s", device_id, "shell", "am", "force-stop", PACKAGE_NAME])
        time.sleep(2)
        subprocess.run(
            [
                "adb",
                "-s",
                device_id,
                "shell",
                "am",
                "start",
                "-n",
                "com.xunmeng.pinduoduo/com.xunmeng.pinduoduo.ui.activity.MainFrameActivity",
            ]
        )
        time.sleep(8)


def search_product(d, keyword: str) -> None:
    print(f"\n搜索：{keyword}")
    width, height = d.window_size()
    search_y = int(height * 200 / 2400)
    d.swipe(width // 2, height // 2, width // 2, height // 2 + 400)
    time.sleep(1)
    d.click(width // 2, search_y)
    time.sleep(0.8)
    edit = d(className="android.widget.EditText")
    if not edit.exists(timeout=3):
        raise RuntimeError("未找到搜索输入框")
    edit.clear_text()
    edit.set_text(keyword)
    time.sleep(0.8)
    btn = d(text="搜索", className="android.widget.TextView")
    if btn.exists(timeout=1):
        btn.click()
    else:
        d.press("enter")
    time.sleep(3)
    try:
        d.wait_idle(timeout=5)
    except Exception:
        pass


def scan_list_products(d, shot_path: str) -> list:
    d.screenshot(shot_path)
    img = cv2.imread(shot_path)
    if img is None:
        return []
    rt = _get_runtime()
    with _infer_lock:
        results = rt["model"](img, conf=0.2)
    products = []
    for r in results:
        for box in r.boxes:
            cls = int(box.cls)
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
            if cls == 0:
                products.append({"type": "baiyi", "cx": cx, "cy": cy})
            elif cls == 1:
                products.append({"type": "brand", "cx": cx, "cy": cy})
            elif cls == 2:
                products.append({"type": "global", "cx": cx, "cy": cy})
    return products


def get_products_with_tags(d) -> list:
    shot_path = f"list_screen_{int(time.time() * 1000)}.jpg"
    prods = scan_list_products(d, shot_path)
    grouped = {}
    for p in prods:
        cx, cy = p["cx"], p["cy"]
        matched = False
        for key in list(grouped.keys()):
            ecx, ecy = key
            if abs(cx - ecx) < 200 and abs(cy - ecy) < 180:
                grouped[key]["tags"].add(p["type"])
                matched = True
                break
        if not matched:
            grouped[(cx, cy)] = {"tags": {p["type"]}, "cx": cx, "cy": cy}
    return list(grouped.values())


def get_priority(tags) -> int:
    if "baiyi" in tags:
        return 4
    if "brand" in tags:
        return 2
    if "global" in tags:
        return 1
    return 0


def is_all_global(item_list) -> bool:
    return all(not ("baiyi" in i["tags"] or "brand" in i["tags"]) for i in item_list)


def scroll_down_once(d) -> None:
    d.swipe(500, 1800, 500, 600, 0.3)
    time.sleep(2.5)


def scroll_to_top(d) -> None:
    width, height = d.window_size()
    for _ in range(2):
        d.swipe(width // 2, height // 2, width // 2, height // 2 + 400)
        time.sleep(0.8)


def sort_products_by_priority(d) -> list:
    raw = get_products_with_tags(d)
    if is_all_global(raw):
        scroll_down_once(d)
        raw = get_products_with_tags(d)
        if is_all_global(raw):
            scroll_down_once(d)
            raw = get_products_with_tags(d)
            if is_all_global(raw):
                scroll_to_top(d)
                raw = get_products_with_tags(d)
    s = sorted(raw, key=lambda x: get_priority(x["tags"]), reverse=True)
    if is_all_global(s):
        s = s[:2]
    return s


def is_subsidy_product(d) -> bool:
    return "百亿补贴" in d.dump_hierarchy() or "官方补贴" in d.dump_hierarchy()


def extract_product_info(xml_content: str, search_word: str) -> Dict[str, Optional[str]]:
    def get_ngram_pairs(text, n=2):
        text = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", text.lower())
        return (
            [text[i : i + n] for i in range(len(text) - n + 1)]
            if len(text) >= n
            else [text]
        )

    def get_single_chars(text):
        text = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", text.lower())
        return [c for c in text]

    def count_chinese(text):
        return len(re.findall(r"[\u4e00-\u9fff]", text))

    search_cn_count = count_chinese(search_word)
    desc_list = re.findall(r'content-desc="([^"]+)"', xml_content)
    best_title = ""
    best_count = 0
    blacklist = [
        "电池",
        "状态栏",
        "电量",
        "百分之",
        "WLAN",
        "手机信号",
        "5G",
        "4G",
        "通知",
        "高德",
        "淘宝",
        "浏览器",
        "手机管家",
        "振铃器",
        "静音",
        "返回",
        "分享",
        "店铺",
        "收藏",
        "客服",
        "工具栏",
        "顶部",
        "拼小圈",
        "¥",
        "￥",
        "大促价",
        "已抢",
        "假一赔十",
        "100%正品",
        "拼单价",
        "狂降",
        "直接成团",
        "买过",
        "次",
        "图片",
        "该店",
        "tronplayer_view",
        "查看全部",
    ]
    search_pairs = get_ngram_pairs(search_word)
    for desc in desc_list:
        desc = desc.strip()
        if any(kw in desc for kw in blacklist):
            continue
        desc_clean = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", desc.lower())
        match_count = sum(1 for p in search_pairs if p in desc_clean)
        if match_count > best_count and match_count > 0:
            best_count = match_count
            best_title = desc
        elif match_count == best_count and match_count > 0:
            if len(desc) > len(best_title):
                best_title = desc
    if not best_title:
        search_chars = get_single_chars(search_word)
        best_count = 0
        for desc in desc_list:
            desc = desc.strip()
            if any(kw in desc for kw in blacklist):
                continue
            if count_chinese(desc) < search_cn_count:
                continue
            desc_clean = re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", desc.lower())
            match_count = sum(1 for c in search_chars if c in desc_clean)
            if match_count > best_count and match_count > 0:
                best_count = match_count
                best_title = desc
            elif match_count == best_count and match_count > 0:
                if len(desc) > len(best_title):
                    best_title = desc
    price_pattern = r"[¥￥]\s*(\d+(?:\.\d+)?)"
    all_prices = re.findall(price_pattern, xml_content)
    price_nums = [float(p) for p in all_prices]
    original_price = None
    current_price = None
    if price_nums:
        prices = sorted(list(set(price_nums)))
        valid = []
        for i in prices:
            keep = True
            for j in prices:
                if i == j:
                    continue
                if max(i, j) >= min(i, j) * 10:
                    s_i = str(int(round(i)))
                    s_j = str(int(round(j)))
                    if len(s_i) >= 3 and len(s_j) >= 3 and s_i[:3] == s_j[:3]:
                        if i > j:
                            keep = False
                        break
            if keep:
                valid.append(i)
        if valid:
            current_price = str(min(valid))
            original_price = str(max(valid))
    return {
        "title": best_title.strip() if best_title else "",
        "original_price": original_price,
        "current_price": current_price,
    }


def find_and_click_detail(d, max_scroll: int = 7) -> bool:
    rt = _get_runtime()
    for _ in range(max_scroll):
        img = d.screenshot(format="opencv")
        with _infer_lock:
            res = rt["detail_model"](img, conf=0.75)
        box = None
        for r in res:
            for b in r.boxes:
                if int(b.cls[0]) == 0:
                    box = tuple(map(int, b.xyxy[0]))
                    break
            if box:
                break
        if box:
            x1, y1, x2, y2 = box
            crop_img = img[y1:y2, x1:x2]
            with _infer_lock:
                ocr_result = rt["reader"].readtext(crop_img)
            full_text = "".join(item[1] for item in ocr_result)
            if "商品详情" in full_text:
                cx = (x1 + x2) // 2
                cy = (y1 + y2) // 2
                d.click(cx, cy)
                time.sleep(1.5)
                if "生产日期" in d.dump_hierarchy():
                    return True
                d.press("back")
                return False
        d.swipe(500, 1800, 500, 600, 0.25)
        time.sleep(0.8)
    return False


def get_date_with_retry(d) -> str:
    m = re.search(r'text="(\d{4}-\d{1,2}-\d{1,2})"', d.dump_hierarchy())
    return m.group(1) if m else ""


def _append_pairs(
    rec_row: List[Any],
    dbg_row: List[Any],
    records_out: Optional[List[List[Any]]],
    debug_out: Optional[List[List[Any]]],
) -> None:
    if records_out is not None:
        records_out.append(rec_row)
        if debug_out is not None:
            debug_out.append(dbg_row)
        return
    with _OUTPUT_LOCK:
        record_list.append(rec_row)
        debug_record_list.append(dbg_row)


def collect_single_product(
    d,
    search_word: str,
    serial_num: int,
    records_out: Optional[List[List[Any]]] = None,
    debug_out: Optional[List[List[Any]]] = None,
) -> Dict[str, Any]:
    xml = d.dump_hierarchy()
    info = extract_product_info(xml, search_word)
    title = info["title"]
    ori = info["original_price"]
    cur = info["current_price"]
    subsidy = "是" if is_subsidy_product(d) else "否"
    detail = find_and_click_detail(d)
    date = get_date_with_retry(d) if detail else ""

    res = validate_product(search_word, title)
    match_pass = bool(res["final"])

    rec_row = [serial_num, title, search_word, ori, cur, subsidy, date]
    if debug_out is not None:
        dbg_n = len(debug_out) + 1
    elif records_out is None:
        dbg_n = len(debug_record_list) + 1
    else:
        dbg_n = 1
    dbg_row = [
        dbg_n,
        search_word,
        title,
        res["s_brand"],
        res["p_brand"],
        "是" if res["brand_ok"] else "否",
        "是" if res["spec_ok"] else "否",
        round(res["ratio"] * 100, 2),
        "是" if res["final"] else "否",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        res["remark"],
    ]
    _append_pairs(rec_row, dbg_row, records_out, debug_out)

    print("\n" + "=" * 80)
    print(f"货品名称：{title}")
    print(f"关键词：{search_word}")
    print(f"原价：{ori} | 现价：{cur}")
    print(f"百亿补贴：{subsidy} | 日期：{date}")
    print(f"校验通过：{match_pass}")
    print("=" * 80)

    return {
        "title": title,
        "subsidy": subsidy,
        "date": date,
        "found_detail": detail,
        "passed": match_pass,
        "remark": res.get("remark", ""),
    }


def select_and_collect_best_product(
    d,
    search_word: str,
    serial_num: int,
    records_out: Optional[List[List[Any]]] = None,
    debug_out: Optional[List[List[Any]]] = None,
) -> Optional[Dict[str, Any]]:
    try:
        scroll_to_top(d)
    except Exception:
        pass

    sorted_prods = sort_products_by_priority(d)
    if not sorted_prods:
        print("未识别商品")
        return None

    priority_order = sorted(
        set(get_priority(p["tags"]) for p in sorted_prods), reverse=True
    )
    any_passed = False
    last_result: Optional[Dict[str, Any]] = None

    for prio in priority_order:
        candidates = [p for p in sorted_prods if get_priority(p["tags"]) == prio]
        if not candidates:
            continue

        if prio == 1 and len(candidates) > 2:
            candidates = candidates[:2]

        print(f"\n===== 优先级 {prio}，共 {len(candidates)} 个商品 =====")
        time.sleep(0.5)

        for i, p in enumerate(candidates):
            print(f"--- 进入商品 {i + 1}/{len(candidates)} ---")
            d.click(p["cx"], p["cy"])
            time.sleep(1)
            res = collect_single_product(d, search_word, serial_num, records_out, debug_out)
            last_result = res
            if res["passed"]:
                any_passed = True
            if res["found_detail"]:
                d.press("back")
                time.sleep(1.5)
            d.press("back")
            time.sleep(1)

        if any_passed:
            print(f"优先级 {prio} 中已有商品通过校验，停止降级")
            break
        print(f"优先级 {prio} 全部未通过，尝试下一优先级")

    if last_result is None:
        return None
    last_result["any_passed"] = any_passed
    return last_result


def collect_one_task(
    device_id: str,
    keyword: str,
    index_num: int,
    *,
    signal: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    _ensure_project_cwd()
    task_records: List[List[Any]] = []
    task_debug: List[List[Any]] = []

    ctx = DeviceContext(device_id)
    d = ctx.d

    _emit(signal, phase="搜索", detail=keyword)
    search_product(d, keyword)

    _emit(signal, phase="列表识别", detail="YOLO / 优先级排序")
    res = select_and_collect_best_product(
        d,
        keyword,
        index_num,
        records_out=task_records,
        debug_out=task_debug,
    )

    _emit(signal, phase="写入汇总", detail="Excel")
    with _OUTPUT_LOCK:
        record_list.extend(task_records)
        debug_record_list.extend(task_debug)
        _write_summary_to_disk()
        _write_debug_to_disk()
        record_list.clear()
        debug_record_list.clear()

    if res is None:
        return {
            "ok": False,
            "passed": False,
            "remark": "未识别商品",
            "status": "未采集",
            "fail_reason": "未识别商品",
        }

    if res.get("any_passed") is True or res.get("passed") is True:
        _emit(signal, phase="完成", detail="已采集")
        return {"ok": True, "passed": True, "remark": "", "status": "已采集"}

    _emit(signal, phase="完成", detail="待复核")
    return {
        "ok": True,
        "passed": False,
        "remark": str(res.get("remark") or "校验未通过"),
        "status": "待复核",
        "fail_reason": "校验未通过",
    }


def _run_standalone_loop(
    task_file: str,
    device_id: str,
    interval_s: int,
) -> None:
    from task_manager import ExcelTaskManager

    mgr = ExcelTaskManager(task_file)
    print("单机顺序领取模式（与 ExcelTaskManager 一致）")
    while True:
        claimed = mgr.claim_next(device_id)
        if claimed is None:
            print("暂无未采集任务或全部在等待重试间隔。")
            break
        idx = claimed.index_num
        try:
            result = collect_one_task(device_id, claimed.keyword, idx)
            status = str(result.get("status") or "未采集")
            remark = str(result.get("remark") or "")
            fr = str(result.get("fail_reason") or "").strip()
            retry_inc = 1 if status == "未采集" else 0
            mgr.finish(
                idx,
                device_id=device_id,
                final_status=status,
                remark=remark,
                retry_inc=retry_inc,
                fail_reason=fr if status == "未采集" else "",
            )
        except Exception as e:
            traceback.print_exc()
            try:
                go_to_pinduoduo_home(device_id)
            except Exception:
                pass
            mgr.finish(
                idx,
                device_id=device_id,
                final_status="未采集",
                remark=f"异常：{str(e)}",
                retry_inc=1,
                fail_reason="运行异常",
            )
        time.sleep(max(0, interval_s))


def main() -> None:
    parser = build_arg_parser()
    args = parser.parse_args()
    _apply_main_args(args)
    standalone_main(args)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser()
    parser.add_argument("--device-id", type=str, default="", help="ADB device serial")
    parser.add_argument("--index-num", type=int, default=0, help="任务序号")
    parser.add_argument("--keyword", type=str, default="", help="搜索词 / 货品名称")
    parser.add_argument("--json", action="store_true", help="输出 JSON 结果到 stdout")
    parser.add_argument("--task-file", type=str, default="", help="任务 Excel 路径")
    parser.add_argument(
        "--interval",
        type=int,
        default=SEARCH_INTERVAL_SECONDS,
        help="单任务完成后等待秒数（单机循环）",
    )
    parser.add_argument(
        "--standalone",
        action="store_true",
        help="单机循环领取未采集任务",
    )
    return parser


def _apply_main_args(args: argparse.Namespace) -> None:
    global PRODUCT_LIST_FILE, SEARCH_INTERVAL_SECONDS
    tf = str(args.task_file or "").strip()
    if tf:
        PRODUCT_LIST_FILE = os.path.abspath(tf)
    if getattr(args, "interval", None) is not None:
        SEARCH_INTERVAL_SECONDS = int(args.interval)


def standalone_main(args: argparse.Namespace) -> None:
    _ensure_project_cwd()

    device_id_out = (
        str(args.device_id).strip()
        if str(args.device_id).strip()
        else os.environ.get("ANDROID_SERIAL", "").strip()
    )

    single_mode = bool(device_id_out and args.index_num and args.keyword)

    if single_mode:
        try:
            result = collect_one_task(device_id_out, args.keyword, int(args.index_num))
            if args.json:
                print(json.dumps(result, ensure_ascii=False))
            else:
                print(result)
            raise SystemExit(0)
        except Exception as e:
            err = {
                "ok": False,
                "status": "未采集",
                "remark": f"异常：{str(e)}",
                "fail_reason": "运行异常",
            }
            if args.json:
                print(json.dumps(err, ensure_ascii=False))
            else:
                print(err)
            raise SystemExit(2)

    if args.standalone:
        if not device_id_out:
            print("未设置设备：请传 --device-id 或环境变量 ANDROID_SERIAL。")
            return
        _run_standalone_loop(PRODUCT_LIST_FILE, device_id_out, int(args.interval or 0))
        print("单机循环结束。")
        return

    if not os.path.exists(PRODUCT_LIST_FILE):
        print("名单文件不存在")
        return

    df = pd.read_excel(PRODUCT_LIST_FILE)
    if "状态" not in df.columns:
        df["状态"] = "未采集"
    if "序号" not in df.columns:
        df["序号"] = range(1, len(df) + 1)
        df.to_excel(PRODUCT_LIST_FILE, index=False)

    todo = df[df["状态"] == "未采集"].copy()
    if todo.empty:
        print("全部完成")
        return

    if not device_id_out:
        print("未设置 ANDROID_SERIAL 或 --device-id，单机模式需要指定设备序列号")
        return

    print("启动采集（兼容模式：按表顺序直接处理未采集行，不经由 claim）")
    for _, row in todo.iterrows():
        kw = str(row["货品名称"]).strip()
        idx = int(row["序号"])
        ok = False
        while not ok:
            try:
                result = collect_one_task(device_id_out, kw, idx)
                df = pd.read_excel(PRODUCT_LIST_FILE)
                df.loc[df["序号"] == idx, "状态"] = result["status"]
                df.to_excel(PRODUCT_LIST_FILE, index=False)
                time.sleep(SEARCH_INTERVAL_SECONDS)
                ok = True
            except Exception as e:
                print(f"异常：{e}")
                traceback.print_exc()
                go_to_pinduoduo_home(device_id_out)
    print("全部采集完成")


if __name__ == "__main__":
    main()
