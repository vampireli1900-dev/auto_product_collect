import os
import subprocess
import sys
import threading
import time
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
import streamlit as st

BASE_DIR = os.path.abspath(
    os.environ.get("COLLECT_PROJECT_DIR") or os.path.dirname(os.path.abspath(__file__))
)
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)
os.chdir(BASE_DIR)

DEFAULT_TASK_PATH = os.path.join(BASE_DIR, "测试用例.xlsx")
LOG_FILE = os.path.join(BASE_DIR, "collector_logs.txt")
COLLECT_SUMMARY_PATH = os.path.join(BASE_DIR, "商品采集汇总.xlsx")

TRACE_MIRROR_HEIGHT = int(os.environ.get("COLLECT_TRACE_MIRROR_HEIGHT", "520"))
TRACE_PREVIEW_HEIGHT = int(os.environ.get("COLLECT_TRACE_PREVIEW_HEIGHT", "320"))

from task_manager import ExcelTaskManager
from main import collect_one_task, go_to_pinduoduo_home

_WORKER_THREADS: Dict[str, threading.Thread] = {}


def list_adb_devices() -> List[str]:
    res = subprocess.check_output(["adb", "devices"], text=True)
    devices: List[str] = []
    for line in res.splitlines():
        if "\tdevice" in line:
            devices.append(line.split("\t")[0])
    return devices


def device_is_online(device_id: str) -> bool:
    try:
        out = subprocess.check_output(
            ["adb", "-s", device_id, "get-state"], text=True
        ).strip()
        return out == "device"
    except Exception:
        return False


def ensure_device_fields(dev: Dict[str, Any]) -> None:
    dev.setdefault("状态", "空闲")
    dev.setdefault("当前任务", "")
    dev.setdefault("已采集数量", 0)
    dev.setdefault("连续失败", 0)
    dev.setdefault("异常信息", "")
    dev.setdefault("阶段", "")
    dev.setdefault("最后活动", "")
    dev.setdefault("采集间隔秒", 40)
    dev.setdefault("每采集休息阈值", 10)
    dev.setdefault("休息秒", 120)
    dev.setdefault("休息到", 0.0)
    dev.setdefault("停止标记", False)
    dev.setdefault("单任务超时秒", 240)
    dev.setdefault("_遥测", {})


def _run_one_collect(
    device_id: str,
    index_num: int,
    keyword: str,
    dev_ref: Dict[str, Any],
    timeout_s: int,
) -> Tuple[Dict[str, Any], str]:
    boxed: Dict[str, Any] = {}
    exc_box: Dict[str, Any] = {}

    tel = dev_ref["_遥测"]
    tel.clear()

    def runner() -> None:
        try:
            boxed["result"] = collect_one_task(
                device_id,
                keyword,
                int(index_num),
                signal=tel,
            )
        except Exception as e:
            exc_box["e"] = e

    th = threading.Thread(target=runner, daemon=True)
    th.start()
    th.join(timeout=float(max(5, timeout_s)))
    if th.is_alive():
        return (
            {"ok": False, "status": "未采集", "remark": f"单任务超时 {timeout_s} 秒"},
            "单任务超时",
        )
    if "e" in exc_box:
        e = exc_box["e"]
        return (
            {
                "ok": False,
                "status": "未采集",
                "remark": f"异常：{str(e)}",
                "fail_reason": "运行异常",
            },
            "运行异常",
        )
    return boxed.get("result") or {"ok": False, "status": "未采集", "remark": "无输出"}, ""


def worker_loop(device_id: str, dev_ref: Dict[str, Any]) -> None:
    task_mgr = ExcelTaskManager(DEFAULT_TASK_PATH)
    ensure_device_fields(dev_ref)
    dev_ref["状态"] = "采集中"
    dev_ref["停止标记"] = False
    dev_ref["异常信息"] = ""
    dev_ref["阶段"] = "初始化"
    dev_ref["最后活动"] = time.strftime("%Y-%m-%d %H:%M:%S")

    while True:
        try:
            dev_ref["最后活动"] = time.strftime("%Y-%m-%d %H:%M:%S")

            if dev_ref.get("停止标记"):
                dev_ref["阶段"] = "已停止"
                dev_ref["状态"] = "空闲"
                dev_ref["当前任务"] = ""
                dev_ref["_遥测"].clear()
                add_log(f"设备 {device_id} 已停止")
                return

            if not device_is_online(device_id):
                dev_ref["阶段"] = "离线检查"
                dev_ref["状态"] = "离线"
                dev_ref["当前任务"] = ""
                dev_ref["_遥测"].clear()
                time.sleep(2)
                continue

            dev_ref["状态"] = "采集中"

            now = time.time()
            rest_until = float(dev_ref.get("休息到") or 0.0)
            if rest_until and now < rest_until:
                dev_ref["阶段"] = "休息中"
                dev_ref["当前任务"] = ""
                time.sleep(1)
                continue

            dev_ref["阶段"] = "抢任务"
            claimed = task_mgr.claim_next(device_id)
            if claimed is None:
                dev_ref["阶段"] = "暂无任务"
                dev_ref["当前任务"] = ""
                time.sleep(1)
                continue

            dev_ref["当前任务"] = f"{claimed.index_num} {claimed.keyword}"
            add_log(f"设备 {device_id} 领取任务：{claimed.index_num} {claimed.keyword}")

            dev_ref["阶段"] = "执行中"
            add_log(f"设备 {device_id} 开始执行：{claimed.index_num} {claimed.keyword}")

            timeout_s = int(dev_ref.get("单任务超时秒") or 240)
            result, forced_fail = _run_one_collect(
                device_id,
                claimed.index_num,
                claimed.keyword,
                dev_ref,
                timeout_s,
            )

            if forced_fail == "单任务超时":
                dev_ref["连续失败"] = int(dev_ref.get("连续失败") or 0) + 1
                dev_ref["异常信息"] = f"单次任务超时 {timeout_s} 秒（设备仍在采集中）"
                add_log(f"设备 {device_id} 超时：{claimed.index_num} {claimed.keyword}")
                task_mgr.finish(
                    claimed.index_num,
                    device_id=device_id,
                    final_status="未采集",
                    remark=f"单任务超时 {timeout_s} 秒",
                    retry_inc=1,
                    fail_reason="超时",
                )
                try:
                    go_to_pinduoduo_home(device_id)
                except Exception:
                    pass
                time.sleep(1)
                continue

            final_status = str(result.get("status") or "未采集")
            remark = str(result.get("remark") or "")
            fr = ""
            if final_status == "未采集":
                fr = str(result.get("fail_reason") or "").strip() or remark or "未采集"
                dev_ref["连续失败"] = int(dev_ref.get("连续失败") or 0) + 1
                dev_ref[
                    "异常信息"
                ] = f"单次失败：{remark or fr}"
            else:
                dev_ref["连续失败"] = 0
                dev_ref["异常信息"] = ""

            task_mgr.finish(
                claimed.index_num,
                device_id=device_id,
                final_status=final_status,
                remark=remark,
                retry_inc=(1 if final_status == "未采集" else 0),
                fail_reason=(fr if final_status == "未采集" else ""),
            )

            add_log(
                f"设备 {device_id} 完成任务：{claimed.index_num} 状态={final_status}"
            )

            if final_status in {"已采集", "待复核"}:
                dev_ref["已采集数量"] = int(dev_ref.get("已采集数量") or 0) + 1

            threshold = int(dev_ref.get("每采集休息阈值") or 0)
            if threshold > 0 and int(dev_ref.get("已采集数量") or 0) % threshold == 0:
                rest_s = int(dev_ref.get("休息秒") or 0)
                if rest_s > 0:
                    dev_ref["休息到"] = time.time() + rest_s
                    add_log(f"设备 {device_id} 进入休息：{rest_s} 秒")

            dev_ref["当前任务"] = ""
            interval_s = int(dev_ref.get("采集间隔秒") or 0)
            if interval_s > 0:
                dev_ref["阶段"] = f"等待间隔 {interval_s} 秒"
                time.sleep(interval_s)
        except Exception as e:
            dev_ref["阶段"] = "异常兜底"
            dev_ref["连续失败"] = int(dev_ref.get("连续失败") or 0) + 1
            dev_ref["异常信息"] = str(e)
            add_log(f"设备 {device_id} worker 异常：{str(e)}")
            time.sleep(1)


st.set_page_config(
    page_title="多机群控采集系统", layout="wide", initial_sidebar_state="expanded"
)


def read_excel_rows(
    path: str, max_rows: int = 2000
) -> Tuple[List[str], List[Dict[str, Any]]]:
    if not os.path.exists(path):
        return [], []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [
            str(x).strip()
            for x in (next(rows) or [])
            if x is not None and str(x).strip() != ""
        ]
    except StopIteration:
        return [], []

    data: List[Dict[str, Any]] = []
    for i, r in enumerate(rows):
        if i >= max_rows:
            break
        if r is None:
            continue
        item: Dict[str, Any] = {}
        for j, col in enumerate(header):
            item[col] = r[j] if j < len(r) else None
        data.append(item)
    return header, data


def read_excel_tail(
    path: str, tail: int = 30
) -> Tuple[List[str], List[Dict[str, Any]]]:
    if not os.path.exists(path):
        return [], []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return [], []
    header_row = values[0] or ()
    header = [
        str(x).strip()
        for x in header_row
        if x is not None and str(x).strip() != ""
    ]
    if not header:
        return [], []
    body = values[1:]
    if tail > 0:
        body = body[-tail:]
    rows: List[Dict[str, Any]] = []
    for r in body:
        item: Dict[str, Any] = {}
        for j, col in enumerate(header):
            item[col] = r[j] if r is not None and j < len(r) else None
        rows.append(item)
    return header, rows


def task_board_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    total = len(rows)
    status_list = [str(r.get("状态") or "").strip() for r in rows]
    captured = sum(1 for s in status_list if s == "已采集")
    review = sum(1 for s in status_list if s == "待复核")
    doing = sum(1 for s in status_list if s == "采集中")
    todo = sum(
        1 for s in status_list if s == "未采集" or s == "" or s.lower() == "nan"
    )
    finished = captured + review
    rate = round(100.0 * captured / (total if total else 1), 2)
    return {
        "total": total,
        "captured": captured,
        "review": review,
        "finished": finished,
        "doing": doing,
        "todo": todo,
        "rate": rate,
    }


def mgr_summary_optional() -> Dict[str, Any]:
    try:
        return ExcelTaskManager(DEFAULT_TASK_PATH).get_summary()
    except Exception:
        return {}


if "task_header" not in st.session_state:
    st.session_state.task_header = []
if "task_rows" not in st.session_state:
    st.session_state.task_rows = []
if "device_list" not in st.session_state:
    st.session_state.device_list = []
if "current_menu" not in st.session_state:
    st.session_state.current_menu = "总览"


def add_log(msg: str) -> None:
    try:
        ts = time.strftime("%Y-%m-%d %H:%M:%S")
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass


def get_recent_logs() -> str:
    try:
        with open(LOG_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()
        return "".join(lines[-100:])
    except Exception:
        return "暂无日志。"


def dedupe_consecutive_lines(text: str) -> str:
    lines = text.rstrip("\n").split("\n")
    out: List[str] = []
    prev = None
    for ln in lines:
        if ln == prev:
            continue
        out.append(ln)
        prev = ln
    return "\n".join(out)


def dedupe_trace_lines(lines: List[str]) -> List[str]:
    out: List[str] = []
    prev = None
    for ln in lines:
        if ln == prev:
            continue
        out.append(ln)
        prev = ln
    return out


def dedupe_fail_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    keys = ("序号", "失败原因", "备注")
    seen = set()
    out_rev: List[Dict[str, Any]] = []
    for r in reversed(rows):
        key = tuple(str(r.get(k) or "").strip() for k in keys)
        if key in seen:
            continue
        seen.add(key)
        out_rev.append(r)
    return list(reversed(out_rev))


def slim_devices_for_table(devs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [{k: v for k, v in d.items() if k != "_遥测"} for d in devs]


def format_trace_for_ui(lines: List[str], tail: int = 600) -> str:
    core = dedupe_trace_lines(lines)
    if tail > 0 and len(core) > tail:
        core = core[-tail:]
    return "\n".join(core)


def get_recent_logs_display() -> str:
    return dedupe_consecutive_lines(get_recent_logs())


def render_trace_mirror(text: str, height_px: int, ui_key: str) -> None:
    st.text_area(
        "采集镜像",
        value=text,
        height=height_px,
        disabled=True,
        label_visibility="collapsed",
        key=ui_key,
    )


with st.sidebar:
    st.markdown("### 多机群控采集")
    st.caption(BASE_DIR)
    st.divider()
    menus = [
        "总览",
        "采集实况",
        "任务管理",
        "设备与调度",
        "采集参数",
        "进度与异常",
        "运行日志",
    ]
    for m in menus:
        if st.button(m, use_container_width=True, key=f"menu_{m}"):
            st.session_state.current_menu = m

menu = st.session_state.current_menu

if menu == "总览":
    st.title("总览")
    ms = mgr_summary_optional()
    if ms:
        total = int(ms.get("total") or 0)
        finished = int(ms.get("finished") or 0)
        todo = int(ms.get("todo") or 0)
        rate = float(ms.get("success_rate_pct") or 0.0)
    else:
        header, rows = read_excel_rows(DEFAULT_TASK_PATH)
        st.session_state.task_header = header
        st.session_state.task_rows = rows
        s = task_board_summary(rows)
        total = s["total"]
        finished = s["finished"]
        todo = s["todo"]
        rate = s["rate"]

    devs = st.session_state.device_list
    online_n = sum(1 for d in devs if device_is_online(d.get("设备ID", "")))

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("总任务数", total)
    col2.metric("已完成（含待复核）", finished)
    col3.metric("待采集", todo)
    col4.metric("整体成功率", f"{rate}%")
    col5.metric("在线设备", online_n)

    st.progress(min(1.0, rate / 100.0) if total > 0 else 0.0)

    st.subheader("设备一览")
    if devs:
        st.dataframe(slim_devices_for_table(devs), use_container_width=True, height=220)
    else:
        st.caption("尚未扫描设备。请到「设备与调度」扫描。")

    with st.expander("最近采集商品", expanded=True):
        _, tail_rows = read_excel_tail(COLLECT_SUMMARY_PATH, tail=30)
        if tail_rows:
            tail_rows = list(reversed(tail_rows))
            st.dataframe(tail_rows, use_container_width=True, height=360)
        else:
            st.caption("暂无采集记录。")

    st.subheader("采集实况预览")
    preview_devs = st.session_state.device_list
    if preview_devs:
        st.caption("完整控制台镜像请看侧边栏「采集实况」。")
        for pdv in preview_devs:
            pid = pdv.get("设备ID") or ""
            plines = (pdv.get("_遥测") or {}).get("采集轨迹") or []
            with st.expander(f"{pid} · 最近输出", expanded=False):
                render_trace_mirror(
                    format_trace_for_ui(plines, tail=400)
                    if plines
                    else "（尚无输出，请先在「设备与调度」扫描并开始采集。）",
                    TRACE_PREVIEW_HEIGHT,
                    ui_key=f"mirror_preview_{pid}",
                )
    else:
        st.caption("尚无设备，请先在「设备与调度」扫描设备。")

elif menu == "采集实况":
    st.title("采集实况（控制台镜像）")
    st.caption(
        "与终端打印一致的优先级流程、商品校验块与汇总分隔线；采集线程写入后即可在此看到。"
        f" 下方窗口高度固定为 {TRACE_MIRROR_HEIGHT}px，内容在框内滚动。"
        " 可用环境变量 COLLECT_TRACE_MIRROR_HEIGHT 调整主窗口像素高度，"
        f"COLLECT_TRACE_PREVIEW_HEIGHT 调整总览预览（默认 {TRACE_PREVIEW_HEIGHT}px）。"
    )
    auto_refresh = st.toggle("自动刷新", value=True)

    devs = st.session_state.device_list
    if not devs:
        st.warning("请先到「设备与调度」扫描设备。")
    else:
        paired = [(d, str(d.get("设备ID") or "")) for d in devs if d.get("设备ID")]
        ids = [p[1] for p in paired]
        tabs = st.tabs(ids)
        for ti, (dev, did) in enumerate(paired):
            ensure_device_fields(dev)
            trace_lines = (dev.get("_遥测") or {}).get("采集轨迹") or []
            tel = dev.get("_遥测") or {}
            head = (
                f"阶段：{tel.get('阶段', '-')}"
                f" · {tel.get('阶段说明', '')}"
                f" · 更新 {tel.get('更新时间', '-')}\n"
                f"当前任务：{dev.get('当前任务') or '-'}\n"
                + ("-" * 72)
                + "\n"
            )
            body = format_trace_for_ui(trace_lines, tail=2500)
            with tabs[ti]:
                render_trace_mirror(
                    head + body if body.strip() else head + "（当前暂无轨迹文本，等待任务执行。）",
                    TRACE_MIRROR_HEIGHT,
                    ui_key=f"mirror_live_{did}",
                )

    if auto_refresh:
        time.sleep(0.8)
        st.rerun()

elif menu == "任务管理":
    st.title("任务列表")
    header, rows = read_excel_rows(DEFAULT_TASK_PATH)
    st.session_state.task_header = header
    st.session_state.task_rows = rows
    st.dataframe(rows, use_container_width=True)

elif menu == "设备与调度":
    st.title("设备与调度")
    auto_refresh = st.toggle("自动刷新", value=True)

    if st.button("扫描设备", use_container_width=True):
        devices = list_adb_devices()
        old = {d.get("设备ID"): d for d in st.session_state.device_list}
        merged = []
        for did in devices:
            if did in old:
                merged.append(old[did])
            else:
                merged.append({"设备ID": did, "状态": "空闲"})
        st.session_state.device_list = merged
        add_log(f"扫描到 {len(devices)} 台设备")
        st.success(f"扫描到 {len(devices)} 台设备。")
        st.rerun()

    st.divider()

    for i, dev in enumerate(st.session_state.device_list):
        device_id = dev["设备ID"]
        ensure_device_fields(dev)
        c1, c2, c3, c4, c5, c6 = st.columns([2, 1, 1, 1, 1, 2])

        with c1:
            online = "在线" if device_is_online(device_id) else "离线"
            st.write(f"**{device_id}** · {online} · {dev.get('状态')}")

        with c2:
            wt = _WORKER_THREADS.get(device_id)
            busy = wt is not None and wt.is_alive()
            label = "开始采集" if not busy else "运行中"
            if st.button(label, key=f"start_{i}", use_container_width=True, disabled=busy):
                dev["停止标记"] = False
                t = threading.Thread(
                    target=worker_loop,
                    args=(device_id, dev),
                    daemon=True,
                )
                _WORKER_THREADS[device_id] = t
                t.start()
                add_log(f"设备 {device_id} 开始采集线程")
                st.rerun()

        with c3:
            if st.button("安全停止", key=f"stop_{i}", use_container_width=True):
                dev["阶段"] = "等待停止"
                dev["停止标记"] = True
                add_log(f"设备 {device_id} 等待当前任务完成后停止")
                st.rerun()

        with c4:
            if st.button("重启 PDD", key=f"restart_{i}", use_container_width=True):
                did = dev["设备ID"]
                subprocess.run(
                    ["adb", "-s", did, "shell", "am force-stop com.xunmeng.pinduoduo"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
                subprocess.run(
                    ["adb", "-s", did, "shell", "am start com.xunmeng.pinduoduo"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                )
                add_log(f"设备 {did} 已重启拼多多")
                st.success(f"{did} 重启命令已下发。")

        with c5:
            st.write("遥测")

        with c6:
            st.write(f"当前任务：{dev.get('当前任务') or '-'}")
            st.write(
                f"本机已累计写入：{dev.get('已采集数量') or 0} · 连续失败：{dev.get('连续失败') or 0}"
            )
            st.write(f"阶段：{dev.get('阶段') or '-'}")
            tel = dev.get("_遥测") or {}
            if tel:
                st.write(
                    f"流水线：{tel.get('阶段', '-')} · {tel.get('阶段说明', '')}"
                    f" · {tel.get('更新时间', '')}"
                )
            st.write(f"最后活动：{dev.get('最后活动') or '-'}")
            if dev.get("异常信息"):
                st.write(f"最近异常（多为单次任务）：{dev.get('异常信息')}")

        trace_lines = (dev.get("_遥测") or {}).get("采集轨迹") or []
        with st.expander(f"实时采集轨迹 · {device_id}", expanded=True):
            render_trace_mirror(
                format_trace_for_ui(trace_lines, tail=1500)
                if trace_lines
                else "（暂无轨迹，开始任务后将在此刷新；也可打开侧边栏「采集实况」。）",
                TRACE_MIRROR_HEIGHT,
                ui_key=f"mirror_sched_{device_id}_{i}",
            )

    st.dataframe(
        slim_devices_for_table(st.session_state.device_list),
        use_container_width=True,
    )
    if auto_refresh:
        time.sleep(1.0)
        st.rerun()

elif menu == "采集参数":
    st.title("采集参数")
    st.caption("按设备生效（运行中时可改，下一任务起作用）。")
    for i, dev in enumerate(st.session_state.device_list):
        ensure_device_fields(dev)
        st.markdown(f"#### {dev.get('设备ID')}")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            dev["采集间隔秒"] = int(
                st.number_input(
                    "采集间隔秒",
                    0,
                    300,
                    int(dev.get("采集间隔秒") or 40),
                    key=f"interval_{i}",
                )
            )
        with c2:
            dev["每采集休息阈值"] = int(
                st.number_input(
                    "每采集 N 条后休息",
                    0,
                    500,
                    int(dev.get("每采集休息阈值") or 10),
                    key=f"rest_n_{i}",
                )
            )
        with c3:
            dev["休息秒"] = int(
                st.number_input(
                    "休息秒",
                    0,
                    3600,
                    int(dev.get("休息秒") or 120),
                    key=f"rest_s_{i}",
                )
            )
        with c4:
            dev["单任务超时秒"] = int(
                st.number_input(
                    "单任务超时秒",
                    30,
                    3600,
                    int(dev.get("单任务超时秒") or 240),
                    key=f"task_to_{i}",
                )
            )

elif menu == "进度与异常":
    st.title("进度与异常")
    auto_refresh = st.toggle("自动刷新", value=True)

    ms = mgr_summary_optional()
    if ms:
        st.metric("成功率（已采集/总任务）", f'{ms.get("success_rate_pct", 0.0)}%')
        st.metric("进行中（表中采集中）", int(ms.get("doing") or 0))

    header, rows = read_excel_rows(DEFAULT_TASK_PATH, max_rows=5000)
    s = task_board_summary(rows)
    st.progress(s["finished"] / s["total"] if s["total"] else 0)

    fail_rows = []
    for r in rows:
        if str(r.get("状态")) == "未采集" and str(r.get("失败原因") or "").strip():
            fail_rows.append(r)
    tail = dedupe_fail_rows(fail_rows)[-50:] if fail_rows else []

    st.subheader("单设备进度")
    st.dataframe(
        slim_devices_for_table(st.session_state.device_list),
        use_container_width=True,
    )
    st.caption(
        "流水线级实时输出请在「采集实况」查看；「设备与调度」内也可展开单设备轨迹。"
    )

    st.subheader("最近失败条目（Excel 中带失败原因的未采集）")
    if tail:
        st.dataframe(list(reversed(tail)), use_container_width=True)
    else:
        st.caption("暂无记录的失败分类。")

    st.subheader("最近采集商品")
    _, tail_rows = read_excel_tail(COLLECT_SUMMARY_PATH, tail=30)
    if tail_rows:
        tail_rows = list(reversed(tail_rows))
        st.dataframe(tail_rows, use_container_width=True, height=280)
    else:
        st.caption("暂无采集记录。")

    if auto_refresh:
        time.sleep(1.0)
        st.rerun()

elif menu == "运行日志":
    st.title("运行日志")
    st.code(get_recent_logs_display())
    time.sleep(1.0)
    st.rerun()
